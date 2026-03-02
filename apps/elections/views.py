"""
ElectON v2 — Elections views.
Includes election CRUD, dashboard, setup wizard, and AJAX helpers.
"""
import json
import logging

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Count, Min, Max, Prefetch, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views import View
from django.views.generic import TemplateView

from apps.audit.decorators import audit_action
from apps.audit.services.audit_service import AuditService
from apps.subscriptions.services import PlanLimitService

from .forms import ElectionForm, PostForm
from .mixins import AjaxRateLimitMixin, ElectionOwnerMixin, PasswordVerifiedMixin
from .models import Election, Post

from apps.candidates.models import Candidate

from apps.voting.models import OFFLINE_VOTER_DOMAIN

logger = logging.getLogger('electon')


def _solana_cluster_param():
    """Return the Solana Explorer cluster query parameter based on SOLANA_NETWORK.

    For localnet, the public Solana Explorer supports ``?cluster=custom&customUrl=<rpc>``.
    """
    network = getattr(settings, 'SOLANA_NETWORK', 'devnet')
    if network == 'localnet':
        rpc_url = getattr(settings, 'SOLANA_RPC_URL', 'http://127.0.0.1:8899')
        return f'cluster=custom&customUrl={rpc_url}'
    elif network == 'mainnet-beta':
        return ''  # mainnet is default, no cluster param needed
    else:
        return f'cluster={network}'


class ManageMyElectionsView(LoginRequiredMixin, TemplateView):
    """List all elections for the current user — with annotated counts (no N+1)."""
    template_name = 'accounts/admin_home.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        elections = (
            Election.objects
            .filter(created_by=self.request.user)
            .annotate(
                total_posts=Count('posts', distinct=True),
                total_candidates=Count('posts__candidates', distinct=True),
                # Exclude revoked voters so stats match dashboard view
                total_voters=Count(
                    'voter_credentials',
                    filter=Q(voter_credentials__is_revoked=False),
                    distinct=True,
                ),
                votes_cast=Count(
                    'voter_credentials',
                    filter=Q(voter_credentials__has_voted=True),
                    distinct=True,
                ),
            )
            .order_by('-created_at')
        )

        context['elections'] = elections

        # ── Aggregate stats (BE-26: DB-level instead of Python memory) ──
        now = timezone.now()
        status_agg = (
            Election.objects
            .filter(created_by=self.request.user)
            .aggregate(
                active_count=Count('id', filter=Q(
                    is_launched=True, start_time__lte=now, end_time__gte=now,
                )),
                concluded_count=Count('id', filter=Q(
                    is_launched=True, end_time__lt=now,
                )),
                pre_launch_count=Count('id', filter=Q(is_launched=False)),
                inactive_count=Count('id', filter=Q(
                    is_launched=True, start_time__gt=now,
                )),
                # Exclude revoked voters from global totals
                total_voters_all=Count(
                    'voter_credentials',
                    filter=Q(voter_credentials__is_revoked=False),
                    distinct=True,
                ),
                total_votes_all=Count(
                    'voter_credentials',
                    filter=Q(voter_credentials__has_voted=True),
                    distinct=True,
                ),
            )
        )
        context['active_count'] = status_agg['active_count']
        context['concluded_count'] = status_agg['concluded_count']
        context['pre_launch_count'] = status_agg['pre_launch_count']
        context['inactive_count'] = status_agg['inactive_count']
        context['total_voters_all'] = status_agg['total_voters_all']
        context['total_votes_all'] = status_agg['total_votes_all']
        user = self.request.user
        context['user'] = user
        # Extract first name from full_name (first word)
        full = (user.full_name or '').strip()
        context['first_name'] = full.split()[0] if full else user.username
        return context


class CreateElectionView(LoginRequiredMixin, View):
    """Create a new election."""
    template_name = 'accounts/create_election.html'

    def get(self, request):
        # Check election limit (plan-based)
        allowed, info = PlanLimitService.check_election_limit(request.user)
        if not allowed:
            messages.error(
                request,
                f'You\'ve reached the maximum of {info["limit"]} elections '
                f'for your {info["plan_name"]} plan.',
            )
            return redirect('elections:manage')

        return render(request, self.template_name, {'form': ElectionForm()})

    @audit_action('election_create')
    def post(self, request):
        form = ElectionForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})

        election = form.save(commit=False)
        election.created_by = request.user
        election.save()

        messages.success(request, f'Election "{election.name}" created successfully.')
        return redirect('elections:dashboard', election_uuid=election.election_uuid)


class EditElectionView(ElectionOwnerMixin, View):
    """Edit election settings (only allowed before launch)."""
    template_name = 'accounts/create_election.html'

    def get(self, request, election_uuid):
        election = self.get_election(election_uuid)
        if not election.can_edit:
            messages.error(request, 'Cannot edit a launched or ended election.')
            return redirect('elections:dashboard', election_uuid=election_uuid)

        form = ElectionForm(instance=election)
        return render(request, self.template_name, {
            'form': form,
            'election': election,
            'is_edit': True,
        })

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)
        if not election.can_edit:
            messages.error(request, 'Cannot edit a launched or ended election.')
            return redirect('elections:dashboard', election_uuid=election_uuid)

        form = ElectionForm(request.POST, instance=election)
        if not form.is_valid():
            return render(request, self.template_name, {
                'form': form,
                'election': election,
                'is_edit': True,
            })

        form.save()
        AuditService.log('election_update', request=request, election=election)
        messages.success(request, f'Election "{election.name}" updated successfully.')
        return redirect('elections:dashboard', election_uuid=election.election_uuid)


class ElectionDashboardView(ElectionOwnerMixin, View):
    """Detailed election dashboard — 6-tab interface with annotated queries."""
    template_name = 'elections/dashboard.html'

    def get(self, request, election_uuid):
        election = self.get_election(election_uuid)

        posts = (
            Post.objects
            .filter(election=election)
            .prefetch_related(
                Prefetch(
                    'candidates',
                    queryset=Candidate.objects.annotate(
                        vote_count=Count('votes')
                    ).order_by('order', 'created_at'),
                ),
            )
            .annotate(
                candidate_count=Count('candidates', distinct=True),
            )
            .order_by('order', 'created_at')
        )

        voter_stats = (
            election.voter_credentials
            .aggregate(
                # Only count active (non-revoked) voters in all totals
                total=Count('id', filter=Q(is_revoked=False)),
                # BUG-06: exclude revoked voters from vote count (prevents turnout > 100%)
                voted=Count('id', filter=Q(has_voted=True, is_revoked=False)),
                invited=Count('id', filter=Q(invitation_sent=True, has_voted=False, is_revoked=False)),
                failed=Count('id', filter=Q(invitation_sent=False, is_revoked=False, has_voted=False) & ~Q(invitation_error='')),
                revoked=Count('id', filter=Q(is_revoked=True)),
                # Offline = emails ending with OFFLINE_VOTER_DOMAIN (non-revoked)
                offline_total=Count('id', filter=Q(voter_email__endswith=OFFLINE_VOTER_DOMAIN, is_revoked=False)),
                offline_voted=Count('id', filter=Q(voter_email__endswith=OFFLINE_VOTER_DOMAIN, has_voted=True, is_revoked=False)),
                # Email-invited = non-offline, non-batch, non-revoked
                email_invited=Count('id', filter=Q(batch_number='') & ~Q(voter_email__endswith=OFFLINE_VOTER_DOMAIN) & Q(is_revoked=False)),
                email_voted=Count('id', filter=Q(batch_number='') & ~Q(voter_email__endswith=OFFLINE_VOTER_DOMAIN) & Q(has_voted=True, is_revoked=False)),
                # PDF credentials = has batch number, non-revoked
                pdf_generated=Count('id', filter=~Q(batch_number='') & Q(is_revoked=False)),
                pdf_voted=Count('id', filter=~Q(batch_number='') & Q(has_voted=True, is_revoked=False)),
            )
        )
        # Derived stats
        voter_stats['email_total'] = voter_stats['total'] - voter_stats['offline_total']
        voter_stats['email_not_voted'] = voter_stats['email_invited'] - voter_stats['email_voted']
        voter_stats['offline_not_voted'] = voter_stats['offline_total'] - voter_stats['offline_voted']
        voter_stats['pdf_not_voted'] = voter_stats['pdf_generated'] - voter_stats['pdf_voted']

        # Batch stats for PDF Generated Credentials table
        voter_batches = list(
            election.voter_credentials
            .exclude(batch_number='')
            .values('batch_number')
            .annotate(
                total=Count('id'),
                voted=Count('id', filter=Q(has_voted=True)),
                revoked_count=Count('id', filter=Q(is_revoked=True)),
                generated_at=Min('created_at'),
                batch_revoked_at=Max('revoked_at'),
            )
            .order_by('batch_number')
        )

        # Evaluate posts queryset once for both total_candidates and per-post checks
        posts_list = list(posts)
        total_posts = len(posts_list)
        total_candidates = sum(p.candidate_count for p in posts_list)
        total_voters = voter_stats['total']
        total_votes = voter_stats['voted']

        # Per-post candidate check (no extra DB queries — uses annotated posts_list)
        posts_without_candidates = [p.name for p in posts_list if p.candidate_count == 0]
        all_posts_have_candidates = total_posts > 0 and len(posts_without_candidates) == 0

        # Turnout percentage
        turnout_pct = round((total_votes / total_voters * 100), 1) if total_voters > 0 else 0
        not_voted = total_voters - total_votes
        not_voted_pct = round(100 - turnout_pct, 1)

        # BUG-02: removed Paginator — template has no pagination UI so all voters
        # beyond page 1 were silently invisible. Evaluate the full queryset directly.
        from django.db.models import BooleanField, Case, Value, When
        voter_qs = (
            election.voter_credentials
            .annotate(
                is_offline=Case(
                    When(voter_email__endswith=OFFLINE_VOTER_DOMAIN, then=Value(True)),
                    default=Value(False),
                    output_field=BooleanField(),
                )
            )
            .order_by('-created_at')
            .values(
                'id', 'voter_email', 'voter_name', 'has_voted',
                'invitation_sent', 'invitation_error', 'invitation_error_code',
                'is_revoked', 'one_time_username', 'created_at', 'is_offline',
                'invited_at', 'revoked_at', 'credentials_resent_at',
                'batch_number', 'updated_at',
            )
        )
        voters_all = list(voter_qs)

        # Setup progress (for Basic overview tab)
        # 3 distinct steps: positions, positions all with candidates, voters
        steps_done = sum([
            total_posts > 0,
            all_posts_have_candidates,
            total_voters > 0,
        ])
        can_launch = total_posts > 0 and all_posts_have_candidates and total_voters > 0
        setup_progress = {
            'has_posts': total_posts > 0,
            'has_candidates': total_candidates > 0,
            'all_posts_have_candidates': all_posts_have_candidates,
            'posts_without_candidates': posts_without_candidates,
            'has_voters': total_voters > 0,
            'can_launch': can_launch,
            'percentage': 100 if can_launch else round(steps_done / 3 * 100),
        }

        # Pending access requests count
        from apps.voting.models import VoterAccessRequest
        pending_access_requests = VoterAccessRequest.objects.filter(
            election=election,
            status=VoterAccessRequest.Status.PENDING,
        ).count()

        context = {
            'election': election,
            'posts': posts_list,
            'voter_stats': voter_stats,
            'voters': voters_all,
            # Pre-filtered list: only email-invited voters (no batch, no offline)
            # Used in the voters tab to get correct S.N. without skipping rows.
            'email_voters': [
                v for v in voters_all
                if not v['batch_number'] and not v['voter_email'].endswith(OFFLINE_VOTER_DOMAIN)
            ],
            'voter_batches': voter_batches,
            'post_form': PostForm(),
            # Stats
            'total_posts': total_posts,
            'total_candidates': total_candidates,
            'total_voters': total_voters,
            'total_votes': total_votes,
            'turnout_pct': turnout_pct,
            'not_voted': not_voted,
            'not_voted_pct': not_voted_pct,
            'setup_progress': setup_progress,
            'pending_access_requests': pending_access_requests,
            # Blockchain explorer (Phase 6)
            'SOLANA_EXPLORER_URL': getattr(settings, 'SOLANA_EXPLORER_URL', ''),
            'SOLANA_NETWORK': getattr(settings, 'SOLANA_NETWORK', 'devnet'),
            'SOLANA_CLUSTER_PARAM': _solana_cluster_param(),
            # BUG-07: use reverse() so URL remains correct if routing changes
            'failed_invitations_url': reverse(
                'notifications:failed_invitations',
                kwargs={'election_uuid': election.election_uuid},
            ),
        }
        return render(request, self.template_name, context)

class ElectionStatsView(ElectionOwnerMixin, View):
    """GET: return live stats for the dashboard stat cards."""

    def get(self, request, election_uuid):
        election = self.get_election(election_uuid)
        # BUG-03 + BUG-06: aligned calculation with dashboard view; fixed voted count
        stats = (
            Election.objects
            .filter(pk=election.pk)
            .annotate(
                _total_posts=Count('posts', distinct=True),
                _total_candidates=Count('posts__candidates', distinct=True),
                # BUG-03: count posts that have ≥1 candidate (for setup_pct step 2)
                _posts_with_candidates=Count(
                    'posts',
                    filter=Q(posts__candidates__isnull=False),
                    distinct=True,
                ),
                # Exclude revoked voters from totals (consistent with dashboard view)
                _total_voters=Count(
                    'voter_credentials',
                    filter=Q(voter_credentials__is_revoked=False),
                    distinct=True,
                ),
                # BUG-06: exclude revoked voters from vote count (prevents turnout > 100%)
                _votes_cast=Count(
                    'voter_credentials',
                    filter=Q(
                        voter_credentials__has_voted=True,
                        voter_credentials__is_revoked=False,
                    ),
                    distinct=True,
                ),
                # Voter sub-tab stats (exclude revoked)
                _email_invited=Count(
                    'voter_credentials',
                    filter=(
                        Q(voter_credentials__batch_number='')
                        & ~Q(voter_credentials__voter_email__endswith=OFFLINE_VOTER_DOMAIN)
                        & Q(voter_credentials__is_revoked=False)
                    ),
                    distinct=True,
                ),
                _pdf_generated=Count(
                    'voter_credentials',
                    filter=(
                        ~Q(voter_credentials__batch_number='')
                        & Q(voter_credentials__is_revoked=False)
                    ),
                    distinct=True,
                ),
            )
            .values(
                '_total_posts', '_total_candidates', '_posts_with_candidates',
                '_total_voters', '_votes_cast', '_email_invited', '_pdf_generated',
            )
            .first()
        )
        # BUG-01: replace assert (stripped by -O) with a proper guard
        if not stats:
            return JsonResponse(
                {'success': False, 'error': 'Election data unavailable.'},
                status=500,
            )
        total_posts = stats['_total_posts']
        total_candidates = stats['_total_candidates']
        total_voters = stats['_total_voters']
        total_votes = stats['_votes_cast']
        # BUG-03: step 2 = ALL posts have candidates (mirrors dashboard view logic)
        all_posts_have_cands = (
            total_posts > 0
            and stats['_posts_with_candidates'] == total_posts
        )
        steps_done = sum([total_posts > 0, all_posts_have_cands, total_voters > 0])
        setup_pct  = 100 if election.can_launch else round(steps_done / 3 * 100)
        return JsonResponse({
            'success': True,
            'posts': total_posts,
            'candidates': total_candidates,
            'voters': total_voters,
            'votes': total_votes,
            'setup_pct': setup_pct,
            # BUG-03: send to JS so overview ring uses same logic as server render
            'all_posts_have_candidates': all_posts_have_cands,
            'voter_email_invited': stats['_email_invited'],
            'voter_pdf_generated': stats['_pdf_generated'],
            # Current lifecycle status — used by the frontend to detect
            # Pre-launch → Inactive → Active → Concluded transitions without a full page load.
            'status': election.current_status.lower(),
        })


class LaunchElectionView(ElectionOwnerMixin, AjaxRateLimitMixin, PasswordVerifiedMixin, View):
    """Launch an election (validates prerequisites). Requires password confirmation."""
    rate_limit_max = 5
    rate_limit_window = 60

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)

        if not election.is_draft:
            return JsonResponse({'success': False, 'error': 'Election has already been launched.'}, status=400)

        # Password gate — launching is irreversible
        ok, err = self.verify_password(request)
        if not ok:
            return JsonResponse({'success': False, 'error': err}, status=403)

        # Check active election limit (plan-based)
        allowed, info = PlanLimitService.check_active_election_limit(request.user)
        if not allowed:
            return JsonResponse({
                'success': False,
                'error': (
                    f'You\'ve reached the maximum of {info["limit"]} active elections '
                    f'for your {info["plan_name"]} plan. '
                    f'Wait for a current election to conclude or upgrade your plan.'
                ),
            }, status=400)

        # Validate launch prerequisites
        errors = []
        posts = election.posts.prefetch_related('candidates').all()

        if not posts.exists():
            errors.append('Election must have at least one position.')

        for post in posts:
            if not post.candidates.all():  # Use .all() to leverage prefetch_related
                errors.append(f'Position "{post.name}" has no candidates.')

        if not election.voter_credentials.exists():
            errors.append('Election must have at least one registered voter.')

        # Block launch if there are unreviewed access requests
        from apps.voting.models import VoterAccessRequest
        pending_requests = VoterAccessRequest.objects.filter(
            election=election,
            status=VoterAccessRequest.Status.PENDING,
        ).count()
        if pending_requests:
            errors.append(
                f'There {"is" if pending_requests == 1 else "are"} {pending_requests} pending '
                f'access request{"" if pending_requests == 1 else "s"}. '
                f'Please approve or reject all requests before launching.'
            )

        if errors:
            return JsonResponse({'success': False, 'error': ' '.join(errors)}, status=400)

        # Atomically lock the election row to prevent concurrent launches
        with transaction.atomic():
            locked = Election.objects.select_for_update().get(pk=election.pk)
            if not locked.is_draft:
                return JsonResponse({'success': False, 'error': 'Election has already been launched.'}, status=400)

            # Deploy the election to Solana before marking it as launched.
            # If the blockchain deploy fails we roll back the entire transaction so
            # the election stays in draft and can be retried.
            if getattr(settings, 'SOLANA_PROGRAM_ID', ''):
                try:
                    from apps.blockchain.services.program_service import ProgramService
                    svc = ProgramService()
                    svc.deploy_election(locked)
                except Exception as exc:
                    logger.exception(
                        "Solana deploy failed for election %s: %s",
                        locked.election_uuid,
                        exc,
                    )
                    # Give a more specific hint when the validator is unreachable
                    hint = str(exc).lower()
                    connection_keywords = [
                        'connect', 'refused', 'timeout', 'unreachable',
                        'errno', 'connectionerror', 'oserror', 'no route',
                        'network', 'httpx', 'rpc', '111',
                    ]
                    if any(kw in hint for kw in connection_keywords):
                        msg = (
                            'Solana validator is not reachable. '
                            'Ensure the local test-validator is running '
                            '(run dev.sh without --no-chain).'
                        )
                    else:
                        msg = (
                            f'Failed to deploy election to Solana: {type(exc).__name__}. '
                            'Please try again or contact support.'
                        )
                    return JsonResponse(
                        {'success': False, 'error': msg},
                        status=500,
                    )

            locked.is_launched = True
            locked.launch_time = timezone.now()
            locked.save(update_fields=['is_launched', 'launch_time', 'updated_at'])
            election = locked

        AuditService.log('election_launch', request=request, election=election)

        # FEAT-06: Fire webhook
        from apps.notifications.services.webhook_service import WebhookService
        WebhookService.dispatch(election, 'election.launched')

        # SSE: push election-launched event
        from apps.elections.event_emitter import emit_event
        emit_event(election.election_uuid, 'election_update',
                   {'field': 'status', 'value': 'launched'},
                   user_id=request.user.pk)

        return JsonResponse({
            'success': True,
            'message': f'Election "{election.name}" has been launched!',
        })


class UpdateElectionNameView(ElectionOwnerMixin, View):
    """Update election name (draft only)."""

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)
        if not election.can_edit:
            return JsonResponse({'success': False, 'error': 'Cannot edit a launched election.'}, status=400)

        name = request.POST.get('name', '').strip()
        if not name or len(name) > 255:
            return JsonResponse({'success': False, 'error': 'Invalid election name.'}, status=400)

        election.name = name
        election.save(update_fields=['name', 'updated_at'])
        AuditService.log('election_update', request=request, election=election, field='name')

        # SSE: push name change
        from apps.elections.event_emitter import emit_event
        emit_event(election.election_uuid, 'election_update',
                   {'field': 'name', 'value': name},
                   user_id=request.user.pk)

        return JsonResponse({'success': True, 'message': 'Election name updated.'})


class UpdateVotingPeriodView(ElectionOwnerMixin, View):
    """Update voting period. Accepts partial data (start_time, end_time, timezone).

    Allowed states:
    - Draft (not launched): always allowed.
    - Launched-Inactive: allowed so admins can correct wrongly-stored times
      (e.g. created before the timezone conversion fix was deployed).
    - Active / Concluded: blocked — changing times mid-vote would be disruptive.
    """

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)
        status = election.current_status
        if status == Election.STATUS_ACTIVE:
            return JsonResponse(
                {'success': False, 'error': 'Cannot change voting period while the election is active.'},
                status=400,
            )
        if status == Election.STATUS_CONCLUDED:
            return JsonResponse(
                {'success': False, 'error': 'Cannot change voting period of a concluded election.'},
                status=400,
            )

        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        tz_name = request.POST.get('timezone')
        updated_fields = []

        from django.utils.dateparse import parse_datetime
        from django.utils.timezone import is_naive, make_aware
        import zoneinfo

        # Validate the incoming timezone name first (if provided)
        if tz_name:
            try:
                zoneinfo.ZoneInfo(tz_name)
            except (zoneinfo.ZoneInfoNotFoundError, KeyError):
                return JsonResponse({'success': False, 'error': 'Invalid timezone.'}, status=400)
            election.timezone = tz_name
            updated_fields.append('timezone')

        # Use the newly-set timezone (or the existing one) to interpret naive
        # datetime-local strings. The user types times in their election timezone,
        # so we must NOT default to UTC when making them aware.
        effective_tz = zoneinfo.ZoneInfo(tz_name if tz_name else election.timezone or 'UTC')

        if start_time:
            parsed = parse_datetime(start_time)
            if not parsed:
                return JsonResponse({'success': False, 'error': 'Invalid start time format.'}, status=400)
            if is_naive(parsed):
                parsed = make_aware(parsed, effective_tz)
            election.start_time = parsed
            updated_fields.append('start_time')

        if end_time:
            parsed = parse_datetime(end_time)
            if not parsed:
                return JsonResponse({'success': False, 'error': 'Invalid end time format.'}, status=400)
            if is_naive(parsed):
                parsed = make_aware(parsed, effective_tz)
            election.end_time = parsed
            updated_fields.append('end_time')

        # Validate end > start
        if election.end_time <= election.start_time:
            return JsonResponse({'success': False, 'error': 'End time must be after start time.'}, status=400)

        if not updated_fields:
            return JsonResponse({'success': False, 'error': 'No fields to update.'}, status=400)

        updated_fields.append('updated_at')
        election.save(update_fields=updated_fields)
        AuditService.log('election_update', request=request, election=election, field='voting_period')

        # SSE: push period change
        from apps.elections.event_emitter import emit_event
        emit_event(election.election_uuid, 'election_update', {
            'field': 'period',
            'start_time': election.start_time.isoformat(),
            'end_time': election.end_time.isoformat(),
        }, user_id=request.user.pk)

        return JsonResponse({'success': True, 'message': 'Voting period updated.'})


class UpdateAdminMessageView(ElectionOwnerMixin, View):
    """Update admin message (any time)."""

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)
        # MED-04: Prevent editing admin message on concluded elections
        if election.current_status == Election.STATUS_CONCLUDED:
            return JsonResponse({'success': False, 'error': 'Cannot edit a concluded election.'}, status=400)

        message_text = request.POST.get('admin_message', '').strip()

        # SEC-04 / SEC-08: limit length to prevent abuse
        max_length = 2000
        if len(message_text) > max_length:
            return JsonResponse(
                {'success': False, 'error': f'Message too long. Maximum {max_length} characters.'},
                status=400,
            )

        # Sanitize: strip HTML tags to prevent stored XSS
        from django.utils.html import strip_tags
        message_text = strip_tags(message_text)

        election.admin_message = message_text
        election.save(update_fields=['admin_message', 'updated_at'])
        AuditService.log('election_update', request=request, election=election, field='admin_message')
        return JsonResponse({'success': True, 'message': 'Admin message updated.'})


class UpdateAbstainView(ElectionOwnerMixin, View):
    """Toggle allow_abstain setting (draft only)."""

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)
        if not election.can_edit:
            return JsonResponse({'success': False, 'error': 'Cannot modify after launch.'}, status=400)

        election.allow_abstain = not election.allow_abstain
        election.save(update_fields=['allow_abstain', 'updated_at'])

        AuditService.log('election_update', request=request, election=election, field='allow_abstain')

        # SSE: push abstain toggle
        from apps.elections.event_emitter import emit_event
        emit_event(election.election_uuid, 'election_update',
                   {'field': 'allow_abstain', 'value': election.allow_abstain},
                   user_id=request.user.pk)

        return JsonResponse({'success': True, 'allow_abstain': election.allow_abstain})



class DeleteElectionView(ElectionOwnerMixin, AjaxRateLimitMixin, PasswordVerifiedMixin, View):
    """Delete an election. Requires password confirmation."""
    rate_limit_max = 5
    rate_limit_window = 60

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)
        if not election.can_delete:
            return JsonResponse(
                {'success': False, 'error': 'This election is currently active and cannot be deleted.'},
                status=400,
            )

        # Password gate — deletion is permanent
        ok, err = self.verify_password(request)
        if not ok:
            return JsonResponse({'success': False, 'error': err}, status=403)

        # Log before deletion, capturing key data for the audit trail
        AuditService.log(
            'election_delete', request=request, election=election,
            extra={'election_name': election.name, 'election_uuid': str(election.election_uuid)},
        )
        election.delete()
        from django.urls import reverse
        return JsonResponse({
            'success': True,
            'message': 'Election deleted.',
            'redirect': reverse('elections:manage'),
        })


class AddPostView(ElectionOwnerMixin, View):
    """Add a new position/post to an election."""

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)
        if not election.can_edit:
            return JsonResponse({'success': False, 'error': 'Cannot edit a launched election.'}, status=400)

        # Check post limit (plan-based)
        allowed, info = PlanLimitService.check_post_limit(election)
        if not allowed:
            return JsonResponse({
                'success': False,
                'error': (
                    f'Maximum of {info["limit"]} positions allowed '
                    f'for your {info["plan_name"]} plan.'
                ),
            }, status=400)

        form = PostForm(request.POST)
        if not form.is_valid():
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)

        post = form.save(commit=False)
        post.election = election
        post.order = election.posts.count()
        post.save()

        # SSE: push stats update
        from apps.elections.event_emitter import emit_event, build_stats_payload
        emit_event(election.election_uuid, 'stats_update',
                   build_stats_payload(election), user_id=request.user.pk)

        return JsonResponse({
            'success': True,
            'message': f'Position "{post.name}" added.',
            'post_id': post.pk,
            'post_name': post.name,
        })


class AddPostsBulkView(ElectionOwnerMixin, View):
    """Add one or more positions at once (JSON body with posts array)."""

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)
        if not election.can_edit:
            return JsonResponse({'success': False, 'error': 'Cannot edit a launched election.'}, status=400)

        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid request body.'}, status=400)

        entries = data.get('posts', [])
        if not entries or not isinstance(entries, list):
            return JsonResponse({'success': False, 'error': 'No positions provided.'}, status=400)

        # Filter to non-empty names
        clean = []
        for entry in entries:
            name = str(entry.get('name', '')).strip()
            if name:
                clean.append({'name': name})

        if not clean:
            return JsonResponse({'success': False, 'error': 'At least one position name is required.'}, status=400)

        # Check plan limit
        allowed, info = PlanLimitService.check_post_limit(election)
        current = info['current']
        limit = info['limit']
        if current + len(clean) > limit:
            return JsonResponse({
                'success': False,
                'error': (
                    f'Adding {len(clean)} position(s) would exceed the limit '
                    f'of {limit} for your {info["plan_name"]} plan. '
                    f'Currently {current} position(s).'
                ),
            }, status=400)

        # BE-19: Validate ALL names first to prevent partial commit
        created = []
        base_order = current
        validated_forms = []
        for i, item in enumerate(clean):
            form = PostForm({'name': item['name']})
            if not form.is_valid():
                first_error = list(form.errors.values())[0][0]
                return JsonResponse({
                    'success': False,
                    'error': f'Row {i + 1}: {first_error}',
                }, status=400)
            validated_forms.append((form, item))

        with transaction.atomic():
            for i, (form, item) in enumerate(validated_forms):
                post_obj = form.save(commit=False)
                post_obj.election = election
                post_obj.order = base_order + i
                post_obj.save()

                created.append({
                    'post_id': post_obj.pk,
                    'post_name': post_obj.name,
                })

        count = len(created)

        # SSE: push stats update
        from apps.elections.event_emitter import emit_event, build_stats_payload
        emit_event(election.election_uuid, 'stats_update',
                   build_stats_payload(election), user_id=request.user.pk)

        return JsonResponse({
            'success': True,
            'posts': created,
            'message': f'{count} position{"s" if count != 1 else ""} added.',
        })


class DeletePostView(ElectionOwnerMixin, View):
    """Delete a position/post from an election."""

    def post(self, request, election_uuid, post_id):
        election = self.get_election(election_uuid)
        if not election.can_edit:
            return JsonResponse({'success': False, 'error': 'Cannot edit a launched election.'}, status=400)

        post = get_object_or_404(Post, pk=post_id, election=election)
        post.delete()

        # SSE: push stats update
        from apps.elections.event_emitter import emit_event, build_stats_payload
        emit_event(election.election_uuid, 'stats_update',
                   build_stats_payload(election), user_id=request.user.pk)

        return JsonResponse({'success': True, 'message': 'Position deleted.'})


class RenamePostView(ElectionOwnerMixin, View):
    """FEAT-01: Rename a position inline (draft elections only)."""

    def post(self, request, election_uuid, post_id):
        election = self.get_election(election_uuid)
        if not election.can_edit:
            return JsonResponse({'success': False, 'error': 'Cannot edit a launched election.'}, status=400)

        try:
            body = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid JSON.'}, status=400)

        name = (body.get('name') or '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Position name cannot be empty.'}, status=400)
        if len(name) > 255:
            return JsonResponse({'success': False, 'error': 'Name too long (max 255 chars).'}, status=400)

        post_obj = get_object_or_404(Post, pk=post_id, election=election)
        post_obj.name = name
        post_obj.save(update_fields=['name'])

        # SSE: push stats update
        from apps.elections.event_emitter import emit_event, build_stats_payload
        emit_event(election.election_uuid, 'stats_update',
                   build_stats_payload(election), user_id=request.user.pk)

        return JsonResponse({'success': True, 'name': post_obj.name})


class DuplicateElectionView(ElectionOwnerMixin, AjaxRateLimitMixin, View):
    rate_limit_max = 10
    rate_limit_window = 60
    """FEAT-02: Duplicate an election as a new draft (copies posts + candidates)."""

    def post(self, request, election_uuid):
        source = self.get_election(election_uuid)

        # Check election limit (plan-based)
        allowed, info = PlanLimitService.check_election_limit(request.user)
        if not allowed:
            return JsonResponse({
                'success': False,
                'error': (
                    f'You\'ve reached the maximum of {info["limit"]} elections '
                    f'for your {info["plan_name"]} plan.'
                ),
            }, status=400)

        # Pre-check post/candidate limits before duplicating
        source_posts = list(source.posts.prefetch_related('candidates').all().order_by('order'))
        _, post_info = PlanLimitService.check_post_limit(source)
        # The new election will have 0 posts, so check if source post count fits
        if len(source_posts) > post_info['limit']:
            return JsonResponse({
                'success': False,
                'error': (
                    f'Cannot duplicate: source has {len(source_posts)} position(s) '
                    f'but your {post_info["plan_name"]} plan allows max {post_info["limit"]}.'
                ),
            }, status=400)

        for src_post in source_posts:
            cand_count = src_post.candidates.count()
            _, cand_info = PlanLimitService.check_candidate_limit(src_post)
            if cand_count > cand_info['limit']:
                return JsonResponse({
                    'success': False,
                    'error': (
                        f'Cannot duplicate: position "{src_post.name}" has {cand_count} '
                        f'candidate(s) but your plan allows max {cand_info["limit"]}.'
                    ),
                }, status=400)

        from datetime import timedelta
        now = timezone.now()

        # Preserve the original election duration
        original_duration = source.end_time - source.start_time

        # Create the duplicate election (wrapped in transaction for atomicity)
        from apps.candidates.models import Candidate
        with transaction.atomic():
            new_election = Election.objects.create(
                name=f'{source.name[:247]} (Copy)',
                start_time=now + timedelta(days=7),
                end_time=now + timedelta(days=7) + original_duration,
                timezone=source.timezone,
                created_by=request.user,
                allow_voter_results_view=source.allow_voter_results_view,
                allow_abstain=source.allow_abstain,
                admin_message=source.admin_message,
            )

            # Duplicate posts and candidates (without images) — reuse pre-fetched data
            for post in source_posts:
                new_post = Post.objects.create(
                    election=new_election,
                    name=post.name,
                    order=post.order,
                )
                candidates_to_create = [
                    Candidate(
                        election=new_election,
                        post=new_post,
                        name=candidate.name,
                        bio=candidate.bio,
                    )
                    for candidate in post.candidates.all().order_by('name')
                ]
                if candidates_to_create:
                    Candidate.objects.bulk_create(candidates_to_create)

        AuditService.log('election_create', request=request, election=new_election)

        # SSE: notify admin-home so election list refreshes in real-time
        from apps.elections.event_emitter import emit_event, build_stats_payload
        emit_event(new_election.election_uuid, 'stats_update',
                   build_stats_payload(new_election), user_id=request.user.pk)

        messages.success(request, f'Election duplicated as "{new_election.name}".')
        return JsonResponse({
            'success': True,
            'message': f'Election duplicated as "{new_election.name}".',
            'redirect': reverse('elections:dashboard', kwargs={'election_uuid': new_election.election_uuid}),
        })


class PreviewBallotView(ElectionOwnerMixin, View):
    """FEAT-09: Preview the voter ballot interface without creating test voters."""

    def get(self, request, election_uuid):
        election = self.get_election(election_uuid)

        posts = (
            election.posts
            .prefetch_related('candidates')
            .order_by('order', 'created_at')
        )

        # Reuse the ballot template with a preview banner
        return render(request, 'voting/ballot.html', {
            'election': election,
            'voter_name': f'{request.user.first_name or "Admin"} (Preview)',
            'posts': posts,
            'admin_message': election.admin_message,
            'is_preview': True,
        })


# ------------------------------------------------------------------
# Bulk Import
# ------------------------------------------------------------------

class BulkElectionImportView(ElectionOwnerMixin, AjaxRateLimitMixin, View):
    rate_limit_max = 10
    rate_limit_window = 60
    """Import positions, candidates, and voters from a single Excel/CSV file."""

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)

        if not election.can_edit:
            return JsonResponse(
                {'success': False, 'error': 'Cannot modify a launched or ended election.'},
                status=400,
            )

        uploaded = request.FILES.get('election_file')
        if not uploaded:
            return JsonResponse({'success': False, 'error': 'No file uploaded.'}, status=400)

        max_size = 10 * 1024 * 1024
        if uploaded.size > max_size:
            return JsonResponse({'success': False, 'error': 'File exceeds 10 MB limit.'}, status=400)

        ext = uploaded.name.rsplit('.', 1)[-1].lower() if '.' in uploaded.name else ''
        if ext not in ('csv', 'xlsx', 'xls'):
            return JsonResponse(
                {'success': False, 'error': 'Only CSV and Excel files are accepted.'},
                status=400,
            )

        from apps.candidates.services.file_service import FileProcessor
        processor = FileProcessor()

        try:
            if ext == 'csv':
                rows = processor._read_csv(uploaded)
            else:
                rows = processor._read_excel(uploaded)
        except Exception:
            return JsonResponse(
                {'success': False, 'error': 'Could not read the file. Check format.'},
                status=400,
            )

        if not rows:
            return JsonResponse({'success': False, 'error': 'File is empty.'}, status=400)

        columns = {c.lower().strip() for c in rows[0].keys()}

        # Initialize result tracking dict
        result = {
            'positions_created': 0,
            'candidates_created': 0,
            'voters_created': 0,
            'errors': [],
            'success': False,
            'message': '',
        }

        # Normalize column names — _read_csv/_read_excel already normalise spaces→underscores.
        # So we must match normalised keys (e.g. 'candidate_name', not 'candidate name').
        def _get(row, *keys):
            """Return first non-empty value matching any key (handles both space and underscore forms)."""
            for k in keys:
                k_norm = k.lower().strip().replace(' ', '_')
                if k_norm in row:
                    return (row[k_norm] or '').strip()
                # Fallback: exact lower match (for raw/un-normalised callers)
                k_exact = k.lower().strip()
                if k_exact in row:
                    return (row[k_exact] or '').strip()
            return ''

        has_position  = 'position' in columns
        has_candidate = any(c in columns for c in ('candidate', 'candidate_name', 'candidate name'))
        has_email     = 'email' in columns

        # ── Pre-check subscription limits before committing ──
        if has_position and has_candidate:
            # Count unique NEW positions in the file
            unique_import_positions = set()
            for row in rows:
                pn = _get(row, 'position').lower()
                if pn:
                    unique_import_positions.add(pn)
            existing_names = set(
                Post.objects.filter(election=election)
                .values_list('name', flat=True)
            )
            existing_lower = {n.lower() for n in existing_names}
            new_positions_count = sum(1 for p in unique_import_positions if p not in existing_lower)
            current_count = len(existing_names)
            _, info = PlanLimitService.check_post_limit(election)
            limit = info['limit']
            if current_count + new_positions_count > limit:
                return JsonResponse({
                    'success': False,
                    'error': (
                        f'Import failed: adding {new_positions_count} new position(s) would bring the total to '
                        f'{current_count + new_positions_count}, exceeding the {limit}-position limit '
                        f'on your {info["plan_name"]} plan.'
                    ),
                }, status=400)

        elif has_email:
            from apps.voting.models import VoterCredential as VC_check
            unique_emails = set()
            for row in rows:
                em = _get(row, 'email').lower()
                if em:
                    unique_emails.add(em)
            existing_emails_count = election.voter_credentials.count()
            existing_emails_set = set(
                election.voter_credentials.values_list('voter_email', flat=True)
            )
            new_voters_count = sum(1 for e in unique_emails if e not in existing_emails_set)
            _, v_info = PlanLimitService.check_voter_limit(election)
            v_limit = v_info['limit']
            if existing_emails_count + new_voters_count > v_limit:
                return JsonResponse({
                    'success': False,
                    'error': (
                        f'Import failed: adding {new_voters_count} new voter(s) would bring the total to '
                        f'{existing_emails_count + new_voters_count}, exceeding the {v_limit}-voter limit '
                        f'on your {v_info["plan_name"]} plan.'
                    ),
                }, status=400)

        from apps.candidates.models import Candidate
        from apps.voting.models import VoterCredential

        try:
            with transaction.atomic():
                if has_position and has_candidate:
                    post_map = {}
                    for row_num, row in enumerate(rows, 1):
                        pos_name = _get(row, 'position')[:255].strip()
                        # LDB-02: Normalize to title case to avoid near-duplicate positions
                        pos_name = pos_name.title() if pos_name else pos_name
                        cand_name = _get(row, 'candidate name', 'candidate')[:255]
                        bio_raw = _get(row, 'candidate bio (optional)', 'bio')[:500]
                        # SEC-03: Strip HTML tags from imported bios
                        from django.utils.html import strip_tags
                        bio = strip_tags(bio_raw)
                        image_url_raw = _get(row, 'candidate image url (optional)', 'image_url')

                        if not pos_name:
                            result['errors'].append(f'Row {row_num}: position name is required.')
                            continue

                        if pos_name.lower() not in post_map:
                            post_obj, created = Post.objects.get_or_create(
                                election=election,
                                name__iexact=pos_name,
                                defaults={
                                    'name': pos_name,
                                    'order': len(post_map),
                                },
                            )
                            post_map[pos_name.lower()] = post_obj
                            if created:
                                result['positions_created'] += 1

                        if cand_name:
                            post_obj = post_map[pos_name.lower()]
                            existing_cand = Candidate.objects.filter(
                                post=post_obj, name__iexact=cand_name
                            ).first()
                            if not existing_cand:
                                cand = Candidate(
                                    election=election,
                                    post=post_obj,
                                    name=cand_name,
                                    bio=bio,
                                )
                                # Fetch image from URL if provided (SSRF-safe)
                                if image_url_raw:
                                    image_file = processor._fetch_image_safely(image_url_raw)
                                    if image_file:
                                        # Resize + convert to WebP in-memory (0 R2 ops)
                                        from apps.candidates.models import prepare_candidate_image
                                        webp_content, ext = prepare_candidate_image(image_file)
                                        cand.image.save(
                                            f"{cand_name[:40].replace(' ', '_')}{ext}",
                                            webp_content,
                                            save=False,
                                        )
                                    else:
                                        result['errors'].append(
                                            f'Row {row_num}: could not fetch image URL (skipped).'
                                        )
                                cand.save()  # single PUT to R2 (image already processed)
                                result['candidates_created'] += 1

                elif has_email:
                    existing_emails = set(
                        election.voter_credentials
                        .values_list('voter_email', flat=True)
                    )
                    for row_num, row in enumerate(rows, 1):
                        email = _get(row, 'email').lower()
                        name = _get(row, 'name')
                        if not email:
                            result['errors'].append(f'Row {row_num}: email is required.')
                            continue
                        if email in existing_emails:
                            continue
                        VoterCredential.generate_credentials(
                            election=election,
                            voter_email=email,
                            voter_name=name,
                        )
                        existing_emails.add(email)
                        result['voters_created'] += 1
                else:
                    return JsonResponse({
                        'success': False,
                        'error': 'Unrecognized file format. Expected columns: position+candidate or email.',
                    }, status=400)

        except Exception:
            logger.exception('Bulk import failed for election %s', election_uuid)
            return JsonResponse({'success': False, 'error': 'Import failed.'}, status=500)

        result['success'] = True
        result['message'] = (
            f"Imported {result['positions_created']} positions, "
            f"{result['candidates_created']} candidates, "
            f"{result['voters_created']} voters."
        )

        # SSE: push stats update after bulk import
        from apps.elections.event_emitter import emit_event, build_stats_payload
        emit_event(election.election_uuid, 'stats_update',
                   build_stats_payload(election), user_id=request.user.pk)

        return JsonResponse(result)


# ══════════════════════════════════════════════════════════════════
# DASHBOARD API ENDPOINTS
# ══════════════════════════════════════════════════════════════════

class ReorderPostsView(ElectionOwnerMixin, View):
    """Reorder positions via drag-and-drop."""

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)

        if not election.can_edit:
            return JsonResponse({'success': False, 'error': 'Cannot reorder after launch.'}, status=400)

        post_ids = []
        content_type = request.content_type or ''
        if 'application/json' in content_type:
            try:
                data = json.loads(request.body)
                post_ids = data.get('post_ids', [])
            except (json.JSONDecodeError, ValueError):
                return JsonResponse({'success': False, 'error': 'Invalid JSON.'}, status=400)
        else:
            # URL-encoded: getlist handles multiple post_ids values
            post_ids = request.POST.getlist('post_ids')

        # LDB-08: Validate that all post IDs are valid integers and belong to this election
        try:
            post_ids = [int(pid) for pid in post_ids]
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid post ID format.'}, status=400)

        election_post_ids = set(
            election.posts.values_list('pk', flat=True)
        )
        if set(post_ids) != election_post_ids:
            return JsonResponse(
                {'success': False, 'error': 'Post IDs do not match this election\'s positions.'},
                status=400,
            )

        # Batch update all post orders in one query (was N+1)
        from django.db.models import Case, When, Value, IntegerField
        cases = [When(pk=pid, then=Value(idx)) for idx, pid in enumerate(post_ids)]
        election.posts.update(order=Case(*cases, output_field=IntegerField()))

        # SSE: push stats update (order changed)
        from apps.elections.event_emitter import emit_event, build_stats_payload
        emit_event(election.election_uuid, 'stats_update',
                   build_stats_payload(election), user_id=request.user.pk)

        return JsonResponse({'success': True, 'message': 'Positions reordered.'})


class ExportTemplateView(ElectionOwnerMixin, View):
    """Download blank CSV/Excel import templates for positions or voters."""

    def get(self, request, election_uuid, template_type):
        election = self.get_election(election_uuid)

        fmt = request.GET.get('format', 'csv').lower()

        if template_type == 'positions':
            if fmt == 'xlsx':
                return self._positions_xlsx()
            return self._positions_csv()
        elif template_type == 'voters':
            if fmt == 'xlsx':
                return self._voters_xlsx()
            return self._voters_csv()
        else:
            return JsonResponse({'error': 'Invalid template type.'}, status=400)

    # ── CSV helpers ──

    @staticmethod
    def _positions_csv():
        import csv as _csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="positions_candidates_template.csv"'
        writer = _csv.writer(response)
        writer.writerow(['Position', 'Candidate Name', 'Candidate Bio (Optional)', 'Candidate Image URL (Optional)'])
        writer.writerow(['President', 'John Doe', 'Brief bio here', ''])
        writer.writerow(['President', 'Jane Smith', 'Another candidate', 'https://example.com/jane.jpg'])
        writer.writerow(['Treasurer', 'Alice Brown', 'Finance background', ''])
        return response

    @staticmethod
    def _voters_csv():
        import csv as _csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="voters_template.csv"'
        writer = _csv.writer(response)
        # Column names intentionally use title-case so they normalise to
        # voter_name / voter_email after k.lower().strip().replace(' ', '_')
        writer.writerow(['S.N.', 'Voter Name', 'Voter Email'])
        writer.writerow([1, 'John Doe', 'john@example.com'])
        writer.writerow([2, 'Jane Smith', 'jane@example.com'])
        return response

    # ── Excel helpers (openpyxl) with field validation rules ──

    @staticmethod
    def _positions_xlsx():
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Protection, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        assert ws is not None  # wb.active is always set on a fresh Workbook
        ws.title = 'Positions & Candidates'

        headers = ['Position', 'Candidate Name', 'Candidate Bio (Optional)', 'Candidate Image URL (Optional)']
        samples = [
            ['President', 'John Doe', 'Brief bio here', ''],
            ['President', 'Jane Smith', 'Another candidate', 'https://example.com/jane.jpg'],
            ['Treasurer', 'Alice Brown', 'Finance background', ''],
        ]
        col_widths = [22, 24, 35, 40]
        total_cols = 4

        # ── Header row ──
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.protection = Protection(locked=True)

        # ── Sample data rows ──
        data_font = Font(name='Calibri', size=11)
        for row_idx, sample in enumerate(samples, 2):
            for col_idx, value in enumerate(sample, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.protection = Protection(locked=False)
                cell.border = thin_border

        # ── Column widths ──
        for col_idx, width in enumerate(col_widths):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

        # ── Protection: lock header row, unlock data cells ──
        ws.protection.sheet = True
        ws.protection.password = ''
        ws.protection.enable()
        for row in range(2, 201):
            for col in range(1, total_cols + 1):
                ws.cell(row=row, column=col).protection = Protection(locked=False)

        # ── Freeze top row ──
        ws.freeze_panes = 'A2'

        # ── Instructions sheet ──
        ins = wb.create_sheet('Instructions')
        ins_rows = [['Column', 'Description', 'Rules']]
        ins_rows.append(['Position', 'Name of the position (e.g. President)', 'Required. Group candidates under the same position name.'])
        ins_rows.append(['Candidate Name', "Candidate's full name", 'Required. Leave blank for positions with no candidates yet.'])
        ins_rows.append(['Candidate Bio (Optional)', 'Short biography or description', 'Optional. Max 500 characters.'])
        ins_rows.append(['Candidate Image URL (Optional)', 'Public URL of candidate photo', 'Optional. Must be https://. JPEG, PNG, GIF, or WebP. Max 5 MB.'])

        ins_header_font = Font(name='Calibri', bold=True, size=11)
        for row_idx, row_data in enumerate(ins_rows, 1):
            for col_idx, val in enumerate(row_data, 1):
                cell = ins.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 1:
                    cell.font = ins_header_font
        ins.column_dimensions['A'].width = 30
        ins.column_dimensions['B'].width = 42
        ins.column_dimensions['C'].width = 62
        ins.protection.sheet = True
        ins.protection.enable()

        # ── Write response ──
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="positions_candidates_template.xlsx"'
        return response

    @staticmethod
    def _voters_xlsx():
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Protection, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        assert ws is not None  # wb.active is always set on a fresh Workbook
        ws.title = 'Voters'

        # Title-case headers normalise to voter_name / voter_email on import
        headers = ['S.N.', 'Voter Name', 'Voter Email']
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.protection = Protection(locked=True)

        samples = [[1, 'John Doe', 'john@example.com'], [2, 'Jane Smith', 'jane@example.com']]
        for row_idx, sample in enumerate(samples, 2):
            for col_idx, val in enumerate(sample, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name='Calibri', size=11)
                cell.border = thin_border
                cell.protection = Protection(locked=False)

        ws.column_dimensions['A'].width = 8   # S.N.
        ws.column_dimensions['B'].width = 28  # Voter Name
        ws.column_dimensions['C'].width = 36  # Voter Email

        # Data validation: Voter Name — max 255 characters
        name_dv = DataValidation(
            type='textLength',
            operator='lessThanOrEqual',
            formula1='255',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle='Name Too Long',
            error='Voter name must be at most 255 characters.',
            showInputMessage=True,
            promptTitle='Voter Name',
            prompt='Full name of the voter. Optional. Max 255 characters.',
        )
        name_dv.add('B2:B1000')
        ws.add_data_validation(name_dv)

        # Data validation: Voter Email — minimum 5 chars (rough guard)
        email_dv = DataValidation(
            type='textLength',
            operator='greaterThanOrEqual',
            formula1='5',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle='Email Required',
            error='Enter a valid email address (e.g. voter@example.com).',
            showInputMessage=True,
            promptTitle='Voter Email',
            prompt='Required. Email address of the voter (e.g. voter@example.com).',
        )
        email_dv.add('C2:C1000')
        ws.add_data_validation(email_dv)

        ws.protection.sheet = True
        ws.protection.password = ''
        ws.protection.enable()
        for row in range(2, 1001):
            for col in range(1, 4):
                ws.cell(row=row, column=col).protection = Protection(locked=False)
        ws.freeze_panes = 'A2'

        # Instructions sheet
        ins = wb.create_sheet('Instructions')
        ins_data = [
            ['Column', 'Description', 'Rules'],
            ['S.N.', 'Serial number', 'Optional. Enter 1, 2, 3\u2026 Ignored during import.'],
            ['Voter Name', "Voter's full name", 'Optional. Max 255 characters.'],
            ['Voter Email', "Voter's email address", 'Required. Must be a valid, unique email address.'],
        ]
        ins_hdr = Font(name='Calibri', bold=True, size=11)
        for r, row_data in enumerate(ins_data, 1):
            for c, val in enumerate(row_data, 1):
                cell = ins.cell(row=r, column=c, value=val)
                if r == 1:
                    cell.font = ins_hdr
        ins.column_dimensions['A'].width = 14
        ins.column_dimensions['B'].width = 25
        ins.column_dimensions['C'].width = 65
        ins.protection.sheet = True
        ins.protection.enable()

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="voters_template.xlsx"'
        return response


class ExportElectionDataView(ElectionOwnerMixin, AjaxRateLimitMixin, View):
    rate_limit_max = 20
    rate_limit_window = 60
    """Export election data (positions+candidates or voters) as CSV."""

    @staticmethod
    def _safe_filename(name):
        """Sanitize election name for use in Content-Disposition filenames."""
        import re as _re
        # Only allow word chars, spaces, hyphens, dots — strip control chars / newlines
        safe = _re.sub(r'[^\w \-.]', '', name).strip()[:80]
        return safe or 'election'

    def _get_voter_status(self, v):
        """Determine voter status string."""
        is_offline = v.voter_email.endswith(OFFLINE_VOTER_DOMAIN)
        if v.is_revoked:
            return 'Revoked', 'Offline' if is_offline else 'Email'
        if v.has_voted:
            return 'Voted', 'Offline' if is_offline else 'Email'
        if is_offline:
            return 'Registered', 'Offline'
        if v.invitation_error:
            return 'Failed', 'Email'
        if v.invitation_sent:
            return 'Invited', 'Email'
        return 'Registered', 'Email'

    def get(self, request, election_uuid):
        import csv

        election = self.get_election(election_uuid)

        export_type = request.GET.get('type', 'all')
        safe_name = self._safe_filename(election.name)
        response = HttpResponse(content_type='text/csv')

        if export_type == 'positions':
            response['Content-Disposition'] = (
                f'attachment; filename="{safe_name}_positions.csv"'
            )
            writer = csv.writer(response)
            writer.writerow(['position', 'candidate', 'bio'])

            posts = (
                Post.objects.filter(election=election)
                .prefetch_related('candidates')
                .order_by('order', 'created_at')
            )
            for post in posts:
                for cand in post.candidates.all():
                    writer.writerow([post.name, cand.name, cand.bio or ''])
                if not post.candidates.exists():
                    writer.writerow([post.name, '', ''])

        elif export_type == 'voters':
            response['Content-Disposition'] = (
                f'attachment; filename="{safe_name}_voters.csv"'
            )
            writer = csv.writer(response)
            writer.writerow(['name', 'email', 'status', 'type'])

            creds = election.voter_credentials.order_by('-created_at')
            for v in creds.iterator():
                status, vtype = self._get_voter_status(v)
                writer.writerow([v.voter_name or '', v.voter_email, status, vtype])

        else:
            # Export all — two separate sections with proper headers
            response['Content-Disposition'] = (
                f'attachment; filename="{safe_name}_data.csv"'
            )
            writer = csv.writer(response)

            # Positions section
            writer.writerow(['section', 'position', 'candidate', 'bio'])
            posts = (
                Post.objects.filter(election=election)
                .prefetch_related('candidates')
                .order_by('order', 'created_at')
            )
            for post in posts:
                for cand in post.candidates.all():
                    writer.writerow(['positions', post.name, cand.name, cand.bio or ''])
                if not post.candidates.exists():
                    writer.writerow(['positions', post.name, '', ''])

            # Blank separator
            writer.writerow([])

            # Voters section
            writer.writerow(['section', 'name', 'email', 'status', 'type'])
            creds = election.voter_credentials.order_by('-created_at')
            for v in creds.iterator():
                status, vtype = self._get_voter_status(v)
                writer.writerow(['voters', v.voter_name or '', v.voter_email, status, vtype])

        # SEC-09: Audit trail for data exports
        AuditService.log(
            'election_export', request=request, election=election,
            field=export_type,
        )

        return response


# ══════════════════════════════════════════════════════════════════
# POSITIONS — DELETE ALL, PDF EXPORT, CANDIDATES JSON
# ══════════════════════════════════════════════════════════════════

class DeleteAllPostsView(ElectionOwnerMixin, AjaxRateLimitMixin, View):
    rate_limit_max = 5
    rate_limit_window = 60
    """Delete every position (and cascade-deletes their candidates)."""

    def post(self, request, election_uuid):
        election = self.get_election(election_uuid)
        if not election.can_edit:
            return JsonResponse(
                {'success': False, 'error': 'Cannot modify a launched election.'},
                status=400,
            )
        count, _ = election.posts.all().delete()
        AuditService.log(
            'election_update', request=request, election=election,
            field='delete_all_posts',
        )

        # SSE: push stats update
        from apps.elections.event_emitter import emit_event, build_stats_payload
        emit_event(election.election_uuid, 'stats_update',
                   build_stats_payload(election), user_id=request.user.pk)

        return JsonResponse({
            'success': True,
            'message': f'{count} position(s) and their candidates deleted.',
        })


class ExportPositionsPDFView(ElectionOwnerMixin, View):
    """Generate a PDF summary of positions + candidates."""

    def get(self, request, election_uuid):
        import io
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )

        election = self.get_election(election_uuid)
        posts = (
            Post.objects.filter(election=election)
            .prefetch_related('candidates')
            .order_by('order', 'created_at')
        )

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
            topMargin=0.75 * inch, bottomMargin=0.75 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'PDFTitle', parent=styles['Heading1'],
            fontSize=16, spaceAfter=6,
        )
        sub_style = ParagraphStyle(
            'PDFSub', parent=styles['Normal'],
            fontSize=9, textColor=colors.grey, spaceAfter=18,
        )
        pos_style = ParagraphStyle(
            'PosName', parent=styles['Heading2'],
            fontSize=12, spaceAfter=4,
        )
        cell_style = ParagraphStyle(
            'Cell', parent=styles['Normal'], fontSize=9,
        )

        story = []
        story.append(Paragraph(f'{election.name} — Positions', title_style))
        story.append(Paragraph(
            f'{posts.count()} position(s) · Generated for election admin',
            sub_style,
        ))

        for post in posts:
            story.append(Paragraph(
                f'{post.order + 1}. {post.name}',
                pos_style,
            ))
            cands = list(post.candidates.all().order_by('name'))
            if cands:
                data: list = [['#', 'Candidate', 'Bio']]
                for i, c in enumerate(cands, 1):
                    bio_text = (c.bio or '—')[:120]
                    data.append([
                        str(i),
                        Paragraph(c.name, cell_style),
                        Paragraph(bio_text, cell_style),
                    ])
                col_widths = [0.4 * inch, 2 * inch, 3.5 * inch]
                tbl = Table(data, colWidths=col_widths, repeatRows=1)
                tbl.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.Color(.94, .94, .96)),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.Color(.3, .3, .35)),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('TOPPADDING', (0, 1), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 0.4, colors.Color(.85, .85, .88)),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                story.append(tbl)
            else:
                story.append(Paragraph(
                    '<font size="9" color="grey">No candidates yet.</font>',
                    styles['Normal'],
                ))
            story.append(Spacer(1, 14))

        doc.build(story)
        buf.seek(0)

        safe_name = ExportElectionDataView._safe_filename(election.name)
        response = HttpResponse(buf.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="{safe_name}_positions.pdf"'
        )
        return response


class GetPostCandidatesView(ElectionOwnerMixin, View):
    """Return candidates for a single position as JSON."""

    def get(self, request, election_uuid, post_id):
        election = self.get_election(election_uuid)
        post_obj = get_object_or_404(Post, pk=post_id, election=election)
        candidates = post_obj.candidates.all().order_by('name')
        data = [
            {
                'id': c.pk,
                'name': c.name,
                'bio': c.bio or '',
                'image_url': c.image.url if c.image else None,
            }
            for c in candidates
        ]
        return JsonResponse({
            'success': True,
            'post_id': post_obj.pk,
            'post_name': post_obj.name,
            'candidates': data,
        })


# ------------------------------------------------------------------
# Access Request Management (admin-facing)
# ------------------------------------------------------------------

class AccessRequestListView(ElectionOwnerMixin, View):
    """GET: return pending access requests for an election as JSON."""

    def get(self, request, election_uuid):
        from apps.voting.models import VoterAccessRequest
        election = self.get_election(election_uuid)
        requests_qs = (
            VoterAccessRequest.objects
            .filter(election=election, status=VoterAccessRequest.Status.PENDING)
            .order_by('-created_at')
            .values('id', 'name', 'email', 'message', 'status', 'created_at', 'reviewed_at')
        )
        data = [
            {
                'id': r['id'],
                'name': r['name'],
                'email': r['email'],
                'message': r['message'],
                'status': r['status'],
                'created_at': r['created_at'].isoformat() if r['created_at'] else None,
                'reviewed_at': r['reviewed_at'].isoformat() if r['reviewed_at'] else None,
            }
            for r in requests_qs
        ]
        return JsonResponse({'success': True, 'requests': data})


class ApproveAccessRequestView(ElectionOwnerMixin, View):
    """POST: approve a pending access request — creates a VoterCredential and sends invitation email."""

    def post(self, request, election_uuid, request_id):
        from apps.voting.models import VoterAccessRequest, VoterCredential
        election = self.get_election(election_uuid)
        access_req = get_object_or_404(
            VoterAccessRequest, pk=request_id, election=election,
        )
        if access_req.status != VoterAccessRequest.Status.PENDING:
            return JsonResponse({
                'success': False,
                'message': f'Request already {access_req.status}.',
            }, status=400)

        # Check if voter already exists
        if VoterCredential.objects.filter(election=election, voter_email=access_req.email).exists():
            access_req.status = VoterAccessRequest.Status.APPROVED
            access_req.reviewed_at = timezone.now()
            access_req.save(update_fields=['status', 'reviewed_at', 'updated_at'])
            return JsonResponse({
                'success': True,
                'message': 'Voter already has credentials. Request marked as approved.',
            })

        # Create voter credential (returns object with _plain_password attached)
        credential = VoterCredential.generate_credentials(
            election=election,
            voter_email=access_req.email,
            voter_name=access_req.name,
        )

        access_req.status = VoterAccessRequest.Status.APPROVED
        access_req.reviewed_at = timezone.now()
        access_req.save(update_fields=['status', 'reviewed_at', 'updated_at'])

        # Send invitation email via Celery task (or synchronous fallback)
        try:
            from apps.notifications.tasks import send_bulk_invitations_task
            send_bulk_invitations_task.delay(  # type: ignore[misc]
                [(credential.pk, credential._plain_password)],  # noqa: SLF001
                election.pk,
            )
        except Exception:
            logger.exception(
                "Failed to queue invitation email for access request %s (election %s)",
                request_id, election_uuid,
            )

        # SSE: push access request update + stats
        from apps.elections.event_emitter import emit_event, build_stats_payload
        emit_event(election.election_uuid, 'access_request',
                   {'action': 'approved', 'request_id': request_id})
        emit_event(election.election_uuid, 'stats_update',
                   build_stats_payload(election), user_id=request.user.pk)

        return JsonResponse({
            'success': True,
            'message': f'Access approved for {access_req.email}. Credentials created and invitation sent.',
            'voter_id': credential.pk,
        })


class RejectAccessRequestView(ElectionOwnerMixin, View):
    """POST: reject a pending access request — deletes the record."""

    def post(self, request, election_uuid, request_id):
        from apps.voting.models import VoterAccessRequest
        election = self.get_election(election_uuid)
        access_req = get_object_or_404(
            VoterAccessRequest, pk=request_id, election=election,
        )
        if access_req.status != VoterAccessRequest.Status.PENDING:
            return JsonResponse({
                'success': False,
                'message': f'Request already {access_req.status}.',
            }, status=400)

        email = access_req.email
        access_req.delete()

        # SSE: push access request rejection
        from apps.elections.event_emitter import emit_event
        emit_event(election.election_uuid, 'access_request',
                   {'action': 'rejected', 'request_id': request_id})

        return JsonResponse({
            'success': True,
            'message': f'Access request from {email} rejected and removed.',
        })
