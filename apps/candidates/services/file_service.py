"""
ElectON v2 — File processing service for import & export.

Handles voter CSV/Excel import/export and candidate CSV/Excel import with
``image_url`` support, SSRF-safe URL fetching, and strict validation.
"""
import csv
import io
import ipaddress
import logging
import os
import re
import zoneinfo
from urllib.parse import urlparse, urlunparse

import pandas as pd
from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = ('.csv', '.xlsx', '.xls')
REQUIRED_COLUMNS = ('voter_name', 'voter_email')

# ── Security: SSRF prevention for image_url fetching ──
_BLOCKED_NETWORKS = [
    ipaddress.ip_network('0.0.0.0/8'),
    ipaddress.ip_network('10.0.0.0/8'),
    ipaddress.ip_network('100.64.0.0/10'),
    ipaddress.ip_network('127.0.0.0/8'),
    ipaddress.ip_network('169.254.0.0/16'),
    ipaddress.ip_network('172.16.0.0/12'),
    ipaddress.ip_network('192.0.0.0/24'),
    ipaddress.ip_network('192.0.2.0/24'),
    ipaddress.ip_network('192.168.0.0/16'),
    ipaddress.ip_network('198.18.0.0/15'),
    ipaddress.ip_network('198.51.100.0/24'),
    ipaddress.ip_network('203.0.113.0/24'),
    ipaddress.ip_network('224.0.0.0/4'),
    ipaddress.ip_network('240.0.0.0/4'),
    ipaddress.ip_network('255.255.255.255/32'),
    ipaddress.ip_network('::1/128'),
    ipaddress.ip_network('fc00::/7'),
    ipaddress.ip_network('fe80::/10'),
    ipaddress.ip_network('ff00::/8'),
]
_IMAGE_FETCH_TIMEOUT = 10   # seconds
_IMAGE_FETCH_MAX_SIZE = 5 * 1024 * 1024  # 5 MB
_ALLOWED_IMAGE_CONTENT_TYPES = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
_MAX_FIELD_LENGTH = 500     # max chars for any text field in import


class FileProcessor:
    """Import / export voters & candidates for an election via CSV or Excel."""

    def __init__(self):
        self.errors: list[str] = []
        self.warnings: list[str] = []

    # ------------------------------------------------------------------
    # Public orchestrators
    # ------------------------------------------------------------------

    def import_voters_from_file(self, uploaded_file, election):
        """
        End-to-end voter import: validate → read → deduplicate → bulk-create.
        """
        max_size = settings.ELECTON_SETTINGS.get('MAX_UPLOAD_SIZE', 5 * 1024 * 1024)

        if uploaded_file.size > max_size:
            return self._fail(f"File too large. Maximum size is {max_size // (1024 * 1024)} MB.")

        filename = uploaded_file.name.lower()
        if not filename.endswith(ALLOWED_EXTENSIONS):
            return self._fail(f"Invalid file format. Allowed: {', '.join(ALLOWED_EXTENSIONS)}.")

        data = self._read_csv(uploaded_file) if filename.endswith('.csv') else self._read_excel(uploaded_file)

        if data is None:
            return self._fail("Unable to read file. Ensure it is a valid CSV/Excel file.")
        if not data:
            return self._fail("The file is empty or contains no data rows.")

        # Row-count cap (plan-aware)
        from apps.subscriptions.services import PlanLimitService
        allowed, info = PlanLimitService.check_import_limit(election.created_by, len(data))
        if not allowed:
            return self._fail(
                f"File contains {len(data)} rows. Maximum allowed is {info['limit']} "
                f"for your {info['plan_name']} plan."
            )

        # Total voter limit (plan-aware)
        allowed, info = PlanLimitService.check_voter_limit(election)
        if not allowed:
            return self._fail(
                f"Election already has {info.get('current', '?')} voters. "
                f"Maximum is {info['limit']} for your {info['plan_name']} plan."
            )

        # BE-39: Combined count check (existing + new rows)
        current_voter_count = info.get('current', 0)
        if current_voter_count + len(data) > info['limit']:
            remaining = info['limit'] - current_voter_count
            return self._fail(
                f"Import would exceed voter limit. Current: {current_voter_count}, "
                f"importing: {len(data)}, limit: {info['limit']}. "
                f"You can import up to {max(0, remaining)} more voters."
            )

        validated = self._validate_voter_rows(data)
        if not validated and self.errors:
            return self._fail("Validation failed.", errors=self.errors)

        unique, duplicates = self._deduplicate_voters(validated, election)

        from apps.voting.models import VoterCredential  # noqa: E402

        created = 0
        try:
            with transaction.atomic():
                for row in unique:
                    VoterCredential.generate_credentials(
                        election=election,
                        voter_email=row['voter_email'],
                        voter_name=row['voter_name'],
                    )
                    created += 1
        except Exception as exc:
            logger.exception("Voter credential creation failed for election %s", election.election_uuid)
            return self._fail("A database error occurred while importing voters. Please try again.")

        msg = f"Import completed. {created} voter(s) imported."
        if duplicates:
            msg += f" {len(duplicates)} duplicate(s) skipped."
        if self.errors:
            msg += f" {len(self.errors)} error(s) encountered."

        return {
            'success': True,
            'message': msg,
            'processed': created,
            'total_rows': len(data),
            'duplicates': len(duplicates),
            'errors': self.errors,
            'warnings': self.warnings,
        }

    def parse_voters_file(self, uploaded_file):
        """
        Parse and validate a voter CSV/Excel file WITHOUT saving anything to the DB.

        Returns a JSON-serialisable dict:
          ``{'success': True, 'voters': [{'email': ..., 'name': ...}, ...], 'count': N}``
        The caller (JS) uses the returned list to populate the manual entry form so
        the user can review before clicking "Send Invitations".
        """
        max_size = settings.ELECTON_SETTINGS.get('MAX_UPLOAD_SIZE', 5 * 1024 * 1024)

        if uploaded_file.size > max_size:
            return self._fail(f"File too large. Maximum size is {max_size // (1024 * 1024)} MB.")

        filename = uploaded_file.name.lower()
        if not filename.endswith(ALLOWED_EXTENSIONS):
            return self._fail(f"Invalid file format. Allowed: {', '.join(ALLOWED_EXTENSIONS)}.")

        data = self._read_csv(uploaded_file) if filename.endswith('.csv') else self._read_excel(uploaded_file)

        if data is None:
            return self._fail("Unable to read file. Ensure it is a valid CSV/Excel file.")
        if not data:
            return self._fail("The file is empty or contains no data rows.")

        validated = self._validate_voter_rows(data)
        if not validated and self.errors:
            return self._fail("Validation failed.", errors=self.errors)

        voters = [{'email': v['voter_email'], 'name': v['voter_name']} for v in validated]
        msg = f"Parsed {len(voters)} voter(s) — review below then click Send Invitations."
        if self.errors:
            msg += f" {len(self.errors)} row(s) had errors and were skipped."

        return {
            'success': True,
            'message': msg,
            'voters': voters,
            'count': len(voters),
            'errors': self.errors,
            'warnings': self.warnings,
        }

    def export_voters_to_csv(self, election):
        """Return an ``HttpResponse`` with voters as CSV.

        Filename: ``{ElectionName}_{YYYYMMDD_HHMMSS}.csv``
        Rows 1-6: election metadata (frozen concept for Excel; plain in CSV).
        Row 7: blank separator.
        Row 8: column headers.
        Row 9+: voter data, naturally sorted (Voter 1, 2, ..., 10 — not alphabetical).
        Username is NOT included.
        """
        from django.db.models import Count, Q
        from django.utils.timezone import localtime
        from apps.voting.models import OFFLINE_VOTER_DOMAIN
        election_tz = zoneinfo.ZoneInfo(election.timezone or 'UTC')

        safe_name = re.sub(r'[^\w\s-]', '', election.name).strip().replace(' ', '_')[:50] or 'election'
        ts = timezone.now().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = (
            f'attachment; filename="{safe_name}_{ts}.csv"'
        )
        writer = csv.writer(response)

        # ── Gather stats ──────────────────────────────────────────────────
        stats = election.voter_credentials.aggregate(
            total=Count('id', filter=Q(is_revoked=False)),
            voted=Count('id', filter=Q(has_voted=True)),
            email_invited=Count(
                'id',
                filter=Q(batch_number='') & ~Q(voter_email__endswith=OFFLINE_VOTER_DOMAIN) & Q(is_revoked=False),
            ),
            inperson=Count('id', filter=Q(voter_email__endswith=OFFLINE_VOTER_DOMAIN)),
        )

        fmt = '%b %d, %Y %H:%M'
        now_str   = localtime(timezone.now(), election_tz).strftime(fmt)
        start_str = localtime(election.start_time, election_tz).strftime(fmt)
        end_str   = localtime(election.end_time, election_tz).strftime(fmt)

        # ── Metadata rows ─────────────────────────────────────────────────
        writer.writerow(['ElectON Voting System'])
        writer.writerow(['Election Name:', election.name])
        writer.writerow(['Election Status:', election.current_status])
        writer.writerow(['Start Time:', start_str])
        writer.writerow(['End Time:', end_str])
        writer.writerow(['Voters Data as of:', now_str])
        writer.writerow([
            'Total Voters:', stats['total'],
            'Email Invited:', stats['email_invited'],
            'In-Person Invited:', stats['inperson'],
            'Total Voted:', stats['voted'],
        ])
        writer.writerow([])  # blank separator

        # ── Column headers ────────────────────────────────────────────────
        writer.writerow([
            'S.N.', 'Name', 'Email', 'Invitation Type',
            'Has Voted', 'Is Revoked',
            'Invited On', 'Revoked On', 'Credentials Resent On',
        ])

        # ── Voter rows — natural sort (Voter 1, 2, ..., 10, not 1, 10, 2) ─
        def _nat_key(vc):
            return [int(c) if c.isdigit() else c.lower()
                    for c in re.split(r'(\d+)', vc.voter_name or '')]

        creds = sorted(election.voter_credentials.all(), key=_nat_key)

        for sn, vc in enumerate(creds, 1):
            is_offline = vc.voter_email.endswith(OFFLINE_VOTER_DOMAIN)
            writer.writerow([
                sn,
                vc.voter_name or '',
                '' if is_offline else vc.voter_email,
                'In-Person' if is_offline else 'Email',
                'YES' if vc.has_voted else 'NO',
                'YES' if vc.is_revoked else 'NO',
                localtime(vc.invited_at, election_tz).strftime(fmt)             if vc.invited_at             else '',
                localtime(vc.revoked_at, election_tz).strftime(fmt)             if vc.revoked_at             else '',
                localtime(vc.credentials_resent_at, election_tz).strftime(fmt)  if vc.credentials_resent_at else '',
            ])
        return response

    def export_voters_to_excel(self, election):
        """Return an ``HttpResponse`` with voters as Excel (.xlsx).

        Rows 1-6: election metadata (frozen via freeze_panes).
        Row 7: blank separator.
        Row 8: column headers.
        Row 9+: voter data, naturally sorted.
        Username is NOT included.
        """
        from django.db.models import Count, Q
        from django.utils.timezone import localtime
        from apps.voting.models import OFFLINE_VOTER_DOMAIN
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        election_tz = zoneinfo.ZoneInfo(election.timezone or 'UTC')

        safe_name = re.sub(r'[^\w\s-]', '', election.name).strip().replace(' ', '_')[:50] or 'election'
        ts = timezone.now().strftime('%Y%m%d_%H%M%S')

        # ── Stats ──────────────────────────────────────────────────────────
        stats = election.voter_credentials.aggregate(
            total=Count('id', filter=Q(is_revoked=False)),
            voted=Count('id', filter=Q(has_voted=True)),
            email_invited=Count(
                'id',
                filter=Q(batch_number='') & ~Q(voter_email__endswith=OFFLINE_VOTER_DOMAIN) & Q(is_revoked=False),
            ),
            inperson=Count('id', filter=Q(voter_email__endswith=OFFLINE_VOTER_DOMAIN)),
        )

        fmt = '%b %d, %Y %H:%M'
        now_str   = localtime(timezone.now(), election_tz).strftime(fmt)
        start_str = localtime(election.start_time, election_tz).strftime(fmt)
        end_str   = localtime(election.end_time, election_tz).strftime(fmt)

        wb = Workbook()
        ws = wb.active
        assert ws is not None  # wb.active is always set on a fresh Workbook()
        ws.title = 'Voters'

        # ── Style helpers ──────────────────────────────────────────────────
        meta_label_font  = Font(name='Calibri', bold=True, size=10)
        meta_value_font  = Font(name='Calibri', size=10)
        branding_font    = Font(name='Calibri', bold=True, size=14, color='007AFF')
        header_font      = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill      = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
        header_align     = Alignment(horizontal='center', vertical='center')
        data_font        = Font(name='Calibri', size=10)
        thin_border      = Border(bottom=Side(style='thin', color='DDDDDD'))
        meta_fill        = PatternFill(start_color='F5F5F7', end_color='F5F5F7', fill_type='solid')

        def set_meta_row(row_num, label, value):
            lbl = ws.cell(row=row_num, column=1, value=label)
            lbl.font = meta_label_font
            lbl.fill = meta_fill
            val = ws.cell(row=row_num, column=2, value=value)
            val.font = meta_value_font
            val.fill = meta_fill

        # ── Branding (row 1) – merge across all 9 data columns ──────────
        brand_cell = ws.cell(row=1, column=1, value='ElectON Voting System')
        brand_cell.font = branding_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        for c in range(2, 10):
            ws.cell(row=1, column=c).fill = meta_fill

        # ── Metadata rows (2-6) – value merges cols 2-9 ───────────────────
        set_meta_row(2, 'Election Name:',     election.name)
        set_meta_row(3, 'Election Status:',   election.current_status)
        set_meta_row(4, 'Start Time:',        start_str)
        set_meta_row(5, 'End Time:',          end_str)
        set_meta_row(6, 'Voters Data as of:', now_str)
        for r in range(2, 7):
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
            for c in range(3, 10):
                ws.cell(row=r, column=c).fill = meta_fill

        # ── Stats row (7) – label/value pairs merged into wider cells ─────
        stat_pairs = [
            ('Total Voters:', stats['total']),
            ('Email Invited:', stats['email_invited']),
            ('In-Person:', stats['inperson']),
            ('Voted:', stats['voted']),
        ]
        col = 1
        for label, value in stat_pairs:
            lbl_cell = ws.cell(row=7, column=col, value=label)
            lbl_cell.font = meta_label_font
            lbl_cell.fill = meta_fill
            val_cell = ws.cell(row=7, column=col + 1, value=value)
            val_cell.font = meta_value_font
            val_cell.fill = meta_fill
            col += 2
        # Fill remaining cells in row 7
        while col <= 9:
            ws.cell(row=7, column=col).fill = meta_fill
            col += 1

        # Row 8: blank separator (still styled with meta fill for visual grouping)
        for col_i in range(1, 10):
            ws.cell(row=8, column=col_i).fill = meta_fill

        # ── Column headers (row 9) ─────────────────────────────────────────
        col_headers = [
            'S.N.', 'Name', 'Email', 'Invitation Type',
            'Has Voted', 'Is Revoked',
            'Invited On', 'Revoked On', 'Credentials Resent On',
        ]
        for col_i, h in enumerate(col_headers, 1):
            cell = ws.cell(row=9, column=col_i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # ── Column widths ──────────────────────────────────────────────────
        col_widths = [6, 26, 32, 16, 12, 12, 20, 20, 24]
        for col_i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_i)].width = w

        # ── Voter data (natural sort) starting at row 10 ────────────────────
        def _nat_key(vc):
            return [int(c) if c.isdigit() else c.lower()
                    for c in re.split(r'(\d+)', vc.voter_name or '')]

        creds = sorted(election.voter_credentials.all(), key=_nat_key)

        for sn, vc in enumerate(creds, 1):
            row_num = sn + 9  # data starts at row 10
            is_offline = vc.voter_email.endswith(OFFLINE_VOTER_DOMAIN)
            row_values = [
                sn,
                vc.voter_name or '',
                '' if is_offline else vc.voter_email,
                'In-Person' if is_offline else 'Email',
                'YES' if vc.has_voted else 'NO',
                'YES' if vc.is_revoked else 'NO',
                localtime(vc.invited_at, election_tz).strftime(fmt)             if vc.invited_at             else '',
                localtime(vc.revoked_at, election_tz).strftime(fmt)             if vc.revoked_at             else '',
                localtime(vc.credentials_resent_at, election_tz).strftime(fmt)  if vc.credentials_resent_at else '',
            ]
            for col_i, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col_i, value=val)
                cell.font = data_font
                cell.border = thin_border

        # ── Freeze rows 1-9 (branding + metadata + headers); data scrolls below ───────
        ws.freeze_panes = 'A10'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="{safe_name}_{ts}.xlsx"'
        )
        return response

    def export_voters_to_pdf(self, election):
        """Return an ``HttpResponse`` with ALL voters as a styled PDF.

        Design:
        - A4 landscape, compact margins.
        - Branded header with election name.
        - Metadata block with election details + stats.
        - Colour-coded data table:
            Green  (#C8E6C9): voted
            Red    (#FFCDD2): revoked
            Yellow (#FFF9C4): invited/pending
            White:            registered (not yet invited)
        - Legend at bottom explaining colour codes.
        - Natural sort (Voter 1, 2, 3 … 10 not 1, 10, 2).
        """
        from django.db.models import Count, Q
        from django.utils.timezone import localtime
        from apps.voting.models import OFFLINE_VOTER_DOMAIN
        election_tz = zoneinfo.ZoneInfo(election.timezone or 'UTC')

        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
        from reportlab.lib.colors import HexColor

        # ── Stats ──────────────────────────────────────────────────────────
        stats = election.voter_credentials.aggregate(
            total=Count('id', filter=Q(is_revoked=False)),
            voted=Count('id', filter=Q(has_voted=True)),
            email_invited=Count(
                'id',
                filter=Q(batch_number='') & ~Q(voter_email__endswith=OFFLINE_VOTER_DOMAIN) & Q(is_revoked=False),
            ),
            inperson=Count('id', filter=Q(voter_email__endswith=OFFLINE_VOTER_DOMAIN)),
        )

        fmt = '%b %d, %Y %H:%M'
        now_str   = localtime(timezone.now(), election_tz).strftime(fmt)
        start_str = localtime(election.start_time, election_tz).strftime(fmt)
        end_str   = localtime(election.end_time, election_tz).strftime(fmt)

        # ── Document setup ─────────────────────────────────────────────────
        buf = io.BytesIO()
        page_size = landscape(A4)
        doc = SimpleDocTemplate(
            buf,
            pagesize=page_size,
            topMargin=0.4 * inch,
            bottomMargin=0.4 * inch,
            leftMargin=0.4 * inch,
            rightMargin=0.4 * inch,
        )
        usable_width = page_size[0] - 0.8 * inch

        styles = getSampleStyleSheet()

        # ── Styles ─────────────────────────────────────────────────────────
        branding_style = ParagraphStyle(
            'Branding', parent=styles['Normal'],
            fontSize=10, textColor=HexColor('#007AFF'), alignment=TA_CENTER, spaceAfter=2,
        )
        title_style = ParagraphStyle(
            'PDFTitle', parent=styles['Heading1'],
            fontSize=15, alignment=TA_CENTER, spaceAfter=4, textColor=HexColor('#1D1D1F'),
        )
        meta_label = ParagraphStyle(
            'MetaLabel', parent=styles['Normal'],
            fontSize=8, textColor=HexColor('#555555'),
        )
        meta_value = ParagraphStyle(
            'MetaValue', parent=styles['Normal'],
            fontSize=8, textColor=HexColor('#1D1D1F'),
        )
        stat_label = ParagraphStyle(
            'StatLabel', parent=styles['Normal'],
            fontSize=8, textColor=HexColor('#555555'), alignment=TA_CENTER,
        )
        stat_value = ParagraphStyle(
            'StatValue', parent=styles['Normal'],
            fontSize=12, textColor=HexColor('#007AFF'), alignment=TA_CENTER,
            fontName='Helvetica-Bold',
        )
        footer_style = ParagraphStyle(
            'Footer', parent=styles['Normal'],
            fontSize=7, textColor=HexColor('#999999'), alignment=TA_RIGHT,
        )

        story = []

        # ── Branding + Title ──────────────────────────────────────────────
        story.append(Paragraph('ElectON Voting System', branding_style))
        story.append(Paragraph(f'Voter List &mdash; {election.name}', title_style))

        # ── Metadata: 2-column key-value ──────────────────────────────────
        info_data = [
            [Paragraph('<b>Election:</b>', meta_label), Paragraph(election.name, meta_value),
             Paragraph('<b>Status:</b>', meta_label), Paragraph(election.current_status, meta_value)],
            [Paragraph('<b>Start:</b>', meta_label), Paragraph(start_str, meta_value),
             Paragraph('<b>End:</b>', meta_label), Paragraph(end_str, meta_value)],
        ]
        info_table = Table(info_data, colWidths=[0.8 * inch, 3.5 * inch, 0.8 * inch, usable_width - 5.1 * inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), HexColor('#F5F5F7')),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROUNDEDCORNERS', [4, 4, 4, 4]),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 6))

        # ── Stats cards (horizontal row) ──────────────────────────────────
        stat_items = [
            ('Total Voters', stats['total']),
            ('Email Invited', stats['email_invited']),
            ('In-Person', stats['inperson']),
            ('Voted', stats['voted']),
        ]
        stats_data = [[Paragraph(f'<b>{v}</b>', stat_value) for _, v in stat_items],
                       [Paragraph(lbl, stat_label) for lbl, _ in stat_items]]
        card_w = usable_width / len(stat_items)
        stats_table = Table(stats_data, colWidths=[card_w] * len(stat_items))
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), HexColor('#F0F4FF')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 6),
            ('TOPPADDING', (0, -1), (-1, -1), 0),
            ('GRID', (0, 0), (-1, -1), 0.3, HexColor('#D8E2F2')),
            ('ROUNDEDCORNERS', [4, 4, 4, 4]),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 8))

        # ── Voter data: natural sort ───────────────────────────────────────
        def _nat_key(vc):
            return [int(c) if c.isdigit() else c.lower()
                    for c in re.split(r'(\d+)', vc.voter_name or '')]

        creds = sorted(election.voter_credentials.all(), key=_nat_key)

        # ── Colour mapping ─────────────────────────────────────────────────
        COLOR_VOTED   = HexColor('#C8E6C9')
        COLOR_REVOKED = HexColor('#FFCDD2')
        COLOR_INVITED = HexColor('#FFF9C4')
        COLOR_WHITE   = HexColor('#FFFFFF')

        def _row_color(vc):
            if vc.has_voted:
                return COLOR_VOTED
            if vc.is_revoked:
                return COLOR_REVOKED
            if vc.invitation_sent or vc.voter_email.endswith(OFFLINE_VOTER_DOMAIN):
                return COLOR_INVITED
            return COLOR_WHITE

        # ── Column widths ─────────────────────────────────────────────────
        col_widths = [
            0.35 * inch,   # S.N.
            1.55 * inch,   # Name
            2.00 * inch,   # Email / Batch
            0.80 * inch,   # Type
            0.65 * inch,   # Voted
            0.70 * inch,   # Revoked
            1.28 * inch,   # Invited On
            1.28 * inch,   # Revoked On
            1.28 * inch,   # Resent On
        ]

        header_style = ParagraphStyle(
            'HdrCell', parent=styles['Normal'],
            fontSize=7.5, textColor=HexColor('#FFFFFF'), alignment=TA_CENTER,
        )
        cell_style = ParagraphStyle(
            'DataCell', parent=styles['Normal'],
            fontSize=7.5, textColor=HexColor('#1D1D1F'),
        )
        cell_center = ParagraphStyle(
            'DataCellCenter', parent=styles['Normal'],
            fontSize=7.5, textColor=HexColor('#1D1D1F'), alignment=TA_CENTER,
        )

        col_headers = [
            Paragraph('S.N.', header_style),
            Paragraph('Name', header_style),
            Paragraph('Email / Batch', header_style),
            Paragraph('Type', header_style),
            Paragraph('Voted', header_style),
            Paragraph('Revoked', header_style),
            Paragraph('Invited On', header_style),
            Paragraph('Revoked On', header_style),
            Paragraph('Resent On', header_style),
        ]

        table_data = [col_headers]
        row_colors = []

        for sn, vc in enumerate(creds, 1):
            is_offline = vc.voter_email.endswith(OFFLINE_VOTER_DOMAIN)
            email_or_batch = vc.batch_number if is_offline else vc.voter_email
            if is_offline:
                invited_on = localtime(vc.created_at, election_tz).strftime(fmt) if vc.created_at else ''
            else:
                invited_on = localtime(vc.invited_at, election_tz).strftime(fmt) if vc.invited_at else ''
            revoked_on = localtime(vc.revoked_at, election_tz).strftime(fmt) if vc.revoked_at else ''
            resent_on  = localtime(vc.credentials_resent_at, election_tz).strftime(fmt) if vc.credentials_resent_at else ''

            row = [
                Paragraph(str(sn), cell_center),
                Paragraph(vc.voter_name or '\u2014', cell_style),
                Paragraph(email_or_batch or '\u2014', cell_style),
                Paragraph('In-Person' if is_offline else 'Email', cell_center),
                Paragraph('Yes' if vc.has_voted else 'No', cell_center),
                Paragraph('Yes' if vc.is_revoked else 'No', cell_center),
                Paragraph(invited_on, cell_center),
                Paragraph(revoked_on, cell_center),
                Paragraph(resent_on, cell_center),
            ]
            table_data.append(row)
            row_colors.append((sn, _row_color(vc)))

        # ── Table styles ──────────────────────────────────────────────────
        tbl_styles = [
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#007AFF')),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.3, HexColor('#CCCCCC')),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FFFFFF'), HexColor('#FAFAFA')]),
        ]
        for row_idx, bg_color in row_colors:
            tbl_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg_color))

        voter_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        voter_table.setStyle(TableStyle(tbl_styles))
        story.append(voter_table)

        # ── Legend ─────────────────────────────────────────────────────────
        story.append(Spacer(1, 8))
        legend_small = ParagraphStyle('LegendSmall', parent=styles['Normal'], fontSize=7, textColor=HexColor('#666666'))
        legend_data = [[
            Paragraph('<b>Legend:</b>', legend_small),
            Paragraph('<font color="#2E7D32">\u25a0</font>  Voted', legend_small),
            Paragraph('<font color="#C62828">\u25a0</font>  Revoked', legend_small),
            Paragraph('<font color="#F57F17">\u25a0</font>  Invited / Generated', legend_small),
            Paragraph('\u25a1  Registered', legend_small),
        ]]
        legend_table = Table(legend_data, colWidths=[0.7 * inch, 0.8 * inch, 0.8 * inch, 1.5 * inch, 1.0 * inch])
        legend_table.setStyle(TableStyle([
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(legend_table)

        # ── Footer ─────────────────────────────────────────────────────────
        story.append(Spacer(1, 4))
        story.append(Paragraph(f'Generated on {now_str}', footer_style))

        # ── Build PDF ──────────────────────────────────────────────────────
        doc.build(story)

        safe_name = re.sub(r'[^\w\s-]', '', election.name).strip().replace(' ', '_')[:50] or 'election'
        ts = timezone.now().strftime('%Y%m%d_%H%M%S')
        buf.seek(0)

        response = HttpResponse(buf.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{safe_name}_{ts}.pdf"'
        return response

    # ═══════════════════════════════════════════════════════════════════
    # ACCESS REQUEST EXPORTS
    # ═══════════════════════════════════════════════════════════════════

    def _get_pending_access_requests(self, election):
        """Return queryset of pending access requests for the election."""
        from apps.voting.models import VoterAccessRequest
        return VoterAccessRequest.objects.filter(
            election=election,
            status=VoterAccessRequest.Status.PENDING,
        ).order_by('-created_at')

    def export_access_requests_to_excel(self, election):
        """Export pending access requests as a styled Excel file."""
        from django.utils.timezone import localtime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        election_tz = zoneinfo.ZoneInfo(election.timezone or 'UTC')
        safe_name = re.sub(r'[^\w\s-]', '', election.name).strip().replace(' ', '_')[:50] or 'election'
        ts = timezone.now().strftime('%Y%m%d_%H%M%S')

        pending_reqs = self._get_pending_access_requests(election)
        fmt = '%b %d, %Y %H:%M'
        now_str = localtime(timezone.now(), election_tz).strftime(fmt)
        start_str = localtime(election.start_time, election_tz).strftime(fmt)
        end_str = localtime(election.end_time, election_tz).strftime(fmt)

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = 'Pending Access Requests'

        # Styles
        branding_font    = Font(name='Calibri', bold=True, size=14, color='007AFF')
        meta_label_font  = Font(name='Calibri', bold=True, size=10)
        meta_value_font  = Font(name='Calibri', size=10)
        header_font      = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill      = PatternFill(start_color='FF9500', end_color='FF9500', fill_type='solid')
        header_align     = Alignment(horizontal='center', vertical='center')
        data_font        = Font(name='Calibri', size=10)
        thin_border      = Border(bottom=Side(style='thin', color='DDDDDD'))
        meta_fill        = PatternFill(start_color='F5F5F7', end_color='F5F5F7', fill_type='solid')

        def set_meta_row(row_num, label, value):
            lbl = ws.cell(row=row_num, column=1, value=label)
            lbl.font = meta_label_font
            lbl.fill = meta_fill
            val = ws.cell(row=row_num, column=2, value=value)
            val.font = meta_value_font
            val.fill = meta_fill

        # Branding – merge across all 4 data columns
        brand_cell = ws.cell(row=1, column=1, value='ElectON Voting System')
        brand_cell.font = branding_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        for c in range(2, 5):
            ws.cell(row=1, column=c).fill = meta_fill

        # Metadata – value merges cols 2-4
        set_meta_row(2, 'Election Name:', election.name)
        set_meta_row(3, 'Election Status:', election.current_status)
        set_meta_row(4, 'Start Time:', start_str)
        set_meta_row(5, 'End Time:', end_str)
        set_meta_row(6, 'Data as of:', now_str)
        for r in range(2, 7):
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            for c in range(3, 5):
                ws.cell(row=r, column=c).fill = meta_fill

        # Stats – merge value across cols 2-4
        lbl = ws.cell(row=7, column=1, value='Pending Requests:')
        lbl.font = meta_label_font
        lbl.fill = meta_fill
        val = ws.cell(row=7, column=2, value=pending_reqs.count())
        val.font = meta_value_font
        val.fill = meta_fill
        ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=4)
        for c in range(3, 5):
            ws.cell(row=7, column=c).fill = meta_fill

        # Blank separator
        for col_i in range(1, 5):
            ws.cell(row=8, column=col_i).fill = meta_fill

        # Column headers (row 9)
        col_headers = ['S.N.', 'Name', 'Email', 'Requested On']
        for col_i, h in enumerate(col_headers, 1):
            cell = ws.cell(row=9, column=col_i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Column widths
        col_widths = [6, 30, 34, 22]
        for col_i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_i)].width = w

        # Data rows
        for sn, req in enumerate(pending_reqs, 1):
            row_num = sn + 9
            requested_on = localtime(req.created_at, election_tz).strftime(fmt) if req.created_at else ''
            row_values = [sn, req.name or '', req.email or '', requested_on]
            for col_i, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col_i, value=val)
                cell.font = data_font
                cell.border = thin_border

        ws.freeze_panes = 'A10'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="Access_Requests_{safe_name}_{ts}.xlsx"'
        )
        return response

    def export_access_requests_to_pdf(self, election):
        """Export pending access requests as a styled PDF (portrait A4)."""
        from django.utils.timezone import localtime
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
        from reportlab.lib.colors import HexColor

        election_tz = zoneinfo.ZoneInfo(election.timezone or 'UTC')
        safe_name = re.sub(r'[^\w\s-]', '', election.name).strip().replace(' ', '_')[:50] or 'election'
        ts = timezone.now().strftime('%Y%m%d_%H%M%S')

        pending_reqs = self._get_pending_access_requests(election)
        req_count = pending_reqs.count()
        fmt = '%b %d, %Y %H:%M'
        now_str   = localtime(timezone.now(), election_tz).strftime(fmt)
        start_str = localtime(election.start_time, election_tz).strftime(fmt)
        end_str   = localtime(election.end_time, election_tz).strftime(fmt)

        # ── Document setup ─────────────────────────────────────────────────
        buf = io.BytesIO()
        page_size = A4
        doc = SimpleDocTemplate(
            buf,
            pagesize=page_size,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
        )
        usable_width = page_size[0] - 1.0 * inch

        styles = getSampleStyleSheet()

        # ── Styles ─────────────────────────────────────────────────────────
        branding_style = ParagraphStyle(
            'Branding', parent=styles['Normal'],
            fontSize=10, textColor=HexColor('#007AFF'), alignment=TA_CENTER, spaceAfter=2,
        )
        title_style = ParagraphStyle(
            'ARTitle', parent=styles['Heading1'],
            fontSize=14, alignment=TA_CENTER, spaceAfter=4, textColor=HexColor('#1D1D1F'),
        )
        meta_label = ParagraphStyle(
            'MetaLabel', parent=styles['Normal'], fontSize=8, textColor=HexColor('#555555'),
        )
        meta_value = ParagraphStyle(
            'MetaValue', parent=styles['Normal'], fontSize=8, textColor=HexColor('#1D1D1F'),
        )
        stat_label = ParagraphStyle(
            'StatLabel', parent=styles['Normal'],
            fontSize=8, textColor=HexColor('#555555'), alignment=TA_CENTER,
        )
        stat_value = ParagraphStyle(
            'StatValue', parent=styles['Normal'],
            fontSize=14, textColor=HexColor('#FF9500'), alignment=TA_CENTER,
            fontName='Helvetica-Bold',
        )
        footer_style = ParagraphStyle(
            'Footer', parent=styles['Normal'],
            fontSize=7, textColor=HexColor('#999999'), alignment=TA_RIGHT,
        )

        story = []

        # ── Branding + Title ──────────────────────────────────────────────
        story.append(Paragraph('ElectON Voting System', branding_style))
        story.append(Paragraph(f'Pending Access Requests &mdash; {election.name}', title_style))

        # ── Metadata: 2-column key-value ──────────────────────────────────
        info_data = [
            [Paragraph('<b>Election:</b>', meta_label), Paragraph(election.name, meta_value),
             Paragraph('<b>Status:</b>', meta_label), Paragraph(election.current_status, meta_value)],
            [Paragraph('<b>Start:</b>', meta_label), Paragraph(start_str, meta_value),
             Paragraph('<b>End:</b>', meta_label), Paragraph(end_str, meta_value)],
        ]
        info_table = Table(info_data, colWidths=[0.8 * inch, 2.3 * inch, 0.8 * inch, usable_width - 3.9 * inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), HexColor('#F5F5F7')),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROUNDEDCORNERS', [4, 4, 4, 4]),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 6))

        # ── Stat card ─────────────────────────────────────────────────────
        stat_data = [
            [Paragraph(f'<b>{req_count}</b>', stat_value)],
            [Paragraph('Pending Requests', stat_label)],
        ]
        stat_table = Table(stat_data, colWidths=[usable_width])
        stat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), HexColor('#FFF8E1')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 6),
            ('TOPPADDING', (0, -1), (-1, -1), 0),
            ('ROUNDEDCORNERS', [4, 4, 4, 4]),
        ]))
        story.append(stat_table)
        story.append(Spacer(1, 8))

        # ── Data table ────────────────────────────────────────────────────
        header_style = ParagraphStyle(
            'HdrCell', parent=styles['Normal'],
            fontSize=9, textColor=HexColor('#FFFFFF'), alignment=TA_CENTER,
        )
        cell_style = ParagraphStyle(
            'DataCell', parent=styles['Normal'],
            fontSize=9, textColor=HexColor('#1D1D1F'),
        )
        cell_center = ParagraphStyle(
            'DataCellCenter', parent=styles['Normal'],
            fontSize=9, textColor=HexColor('#1D1D1F'), alignment=TA_CENTER,
        )

        col_widths_pdf = [0.4 * inch, 2.0 * inch, 2.6 * inch, 1.6 * inch]
        col_headers_pdf = [
            Paragraph('S.N.', header_style),
            Paragraph('Name', header_style),
            Paragraph('Email', header_style),
            Paragraph('Requested On', header_style),
        ]

        table_data = [col_headers_pdf]
        for sn, req in enumerate(pending_reqs, 1):
            requested_on = localtime(req.created_at, election_tz).strftime(fmt) if req.created_at else ''
            table_data.append([
                Paragraph(str(sn), cell_center),
                Paragraph(req.name or '\u2014', cell_style),
                Paragraph(req.email or '\u2014', cell_style),
                Paragraph(requested_on, cell_center),
            ])

        tbl_styles = [
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#FF9500')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.3, HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FFFFFF'), HexColor('#FAFAFA')]),
        ]

        data_table = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
        data_table.setStyle(TableStyle(tbl_styles))
        story.append(data_table)

        # ── Footer ─────────────────────────────────────────────────────────
        story.append(Spacer(1, 6))
        story.append(Paragraph(f'Generated on {now_str}', footer_style))

        # ── Build PDF ──────────────────────────────────────────────────────
        doc.build(story)
        buf.seek(0)

        response = HttpResponse(buf.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Access_Requests_{safe_name}_{ts}.pdf"'
        return response

    def generate_sample_template(self, format_type='csv'):
        """Return a downloadable sample template (CSV or Excel) for voter import."""
        sample = [
            {'voter_name': 'John Doe', 'voter_email': 'john.doe@example.com'},
            {'voter_name': 'Jane Smith', 'voter_email': 'jane.smith@example.com'},
        ]

        if format_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="voter_import_template.csv"'
            writer = csv.DictWriter(response, fieldnames=['voter_name', 'voter_email'])
            writer.writeheader()
            writer.writerows(sample)
            return response

        df = pd.DataFrame(sample)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as w:
            df.to_excel(w, index=False, sheet_name='Voters')
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="voter_import_template.xlsx"'
        return response

    # ------------------------------------------------------------------
    # Candidate import
    # ------------------------------------------------------------------

    CANDIDATE_REQUIRED_COLUMNS = ('name', 'post')
    MAX_CANDIDATES_PER_IMPORT = 200

    def import_candidates_from_file(self, uploaded_file, election):
        """
        Import candidates from CSV/Excel.
        Required columns: ``name``, ``post``
        Optional columns: ``bio``, ``image_url``

        ``post`` must match an existing Post.name in the election (case-insensitive).
        ``image_url`` is an SSRF-safe-fetched image URL (optional).
        """
        max_size = settings.ELECTON_SETTINGS.get('MAX_UPLOAD_SIZE', 5 * 1024 * 1024)
        self.errors = []
        self.warnings = []

        if uploaded_file.size > max_size:
            return self._fail(f"File too large. Maximum size is {max_size // (1024 * 1024)} MB.")

        filename = uploaded_file.name.lower()
        if not filename.endswith(ALLOWED_EXTENSIONS):
            return self._fail(f"Invalid file format. Allowed: {', '.join(ALLOWED_EXTENSIONS)}.")

        data = self._read_csv(uploaded_file) if filename.endswith('.csv') else self._read_excel(uploaded_file)

        if data is None:
            return self._fail("Unable to read file. Ensure it is a valid CSV/Excel file.")
        if not data:
            return self._fail("The file is empty or contains no data rows.")

        if len(data) > self.MAX_CANDIDATES_PER_IMPORT:
            return self._fail(
                f"File contains {len(data)} rows. Maximum allowed is {self.MAX_CANDIDATES_PER_IMPORT}."
            )

        first_row = data[0]
        missing = [c for c in self.CANDIDATE_REQUIRED_COLUMNS if c not in first_row]
        if missing:
            return self._fail(f"Missing required columns: {', '.join(missing)}")

        # Build post lookup (case-insensitive)
        from apps.elections.models import Post
        posts_qs = Post.objects.filter(election=election)
        post_map = {p.name.strip().lower(): p for p in posts_qs}

        if not post_map:
            return self._fail("Election has no posts. Create posts before importing candidates.")

        # Validate & collect
        from apps.candidates.models import Candidate
        validated = []
        seen = set()

        for idx, row in enumerate(data, start=1):
            name = self._sanitize_field(row.get('name', ''), 'name', idx)
            post_name = self._sanitize_field(row.get('post', ''), 'post', idx)
            bio = self._sanitize_field(row.get('bio', ''), 'bio', idx, required=False)
            image_url_raw = row.get('image_url', '').strip()

            if name is None or post_name is None:
                continue

            post_obj = post_map.get(post_name.lower())
            if not post_obj:
                self.errors.append(
                    f"Row {idx}: post '{post_name}' not found. "
                    f"Available: {', '.join(p.name for p in posts_qs)}"
                )
                continue

            # Duplicate within file
            dedup_key = (name.lower(), post_obj.pk)
            if dedup_key in seen:
                self.errors.append(f"Row {idx}: duplicate candidate '{name}' for post '{post_name}'.")
                continue
            seen.add(dedup_key)

            validated.append({
                'name': name,
                'post': post_obj,
                'bio': bio or '',
                'image_url': image_url_raw,
            })

        if not validated and self.errors:
            return self._fail("Validation failed.", errors=self.errors)

        # Deduplicate against existing candidates (BE-38: case-insensitive)
        existing = set(
            (name.lower(), pid)
            for name, pid in Candidate.objects.filter(election=election)
            .values_list('name', 'post_id')
        )
        unique = []
        duplicate_count = 0
        for row in validated:
            if (row['name'].lower(), row['post'].pk) in existing:
                duplicate_count += 1
                self.warnings.append(
                    f"Candidate '{row['name']}' already exists for post '{row['post'].name}'."
                )
            else:
                unique.append(row)

        # Create candidates (use individual save() to trigger _resize_image for images)
        # Plan-based candidate limit check per post (BE-26)
        from apps.subscriptions.services import PlanLimitService
        per_post_count = {}
        for row in unique:
            per_post_count.setdefault(row['post'].pk, 0)
            per_post_count[row['post'].pk] += 1

        for row_post in set(r['post'] for r in unique):
            _, info = PlanLimitService.check_candidate_limit(row_post)
            new_count = per_post_count.get(row_post.pk, 0)
            if info['current'] + new_count > info['limit']:
                return self._fail(
                    f"Adding {new_count} candidate(s) to '{row_post.name}' would exceed "
                    f"the plan limit of {info['limit']} (currently {info['current']}). "
                    f"Reduce the import or upgrade your plan.",
                    errors=self.errors,
                )

        created = 0
        image_errors = 0
        try:
            with transaction.atomic():
                for row in unique:
                    cand = Candidate(
                        election=election,
                        post=row['post'],
                        name=row['name'],
                        bio=row['bio'],
                    )

                    # Fetch image from URL if provided
                    if row['image_url']:
                        image_file = self._fetch_image_safely(row['image_url'])
                        if image_file:
                            # Resize + convert to WebP in-memory (0 R2 ops)
                            from ..models import prepare_candidate_image
                            webp_content, ext = prepare_candidate_image(image_file)
                            cand.image.save(
                                f"{row['name'][:40].replace(' ', '_')}{ext}",
                                webp_content,
                                save=False,
                            )
                        else:
                            image_errors += 1
                            self.warnings.append(
                                f"Could not fetch image for '{row['name']}': URL skipped."
                            )

                    cand.save()  # single PUT to R2 (image already processed)
                    created += 1
        except Exception as exc:
            logger.exception("Candidate import failed for election %s", election.election_uuid)
            return self._fail("A database error occurred while importing candidates. Please try again.")

        msg = f"Import completed. {created} candidate(s) imported."
        if duplicate_count:
            msg += f" {duplicate_count} duplicate(s) skipped."
        if image_errors:
            msg += f" {image_errors} image(s) could not be fetched."
        if self.errors:
            msg += f" {len(self.errors)} error(s) encountered."

        return {
            'success': True,
            'message': msg,
            'processed': created,
            'total_rows': len(data),
            'duplicates': duplicate_count,
            'errors': self.errors,
            'warnings': self.warnings,
        }

    def generate_candidate_template(self, election, format_type='csv'):
        """Return a downloadable sample template for candidate import."""
        from apps.elections.models import Post
        posts = list(Post.objects.filter(election=election).values_list('name', flat=True))
        post_example = posts[0] if posts else 'President'

        sample = [
            {'name': 'Alice Johnson', 'post': post_example, 'bio': 'Experienced leader', 'image_url': ''},
            {'name': 'Bob Williams', 'post': post_example, 'bio': '', 'image_url': 'https://example.com/bob.jpg'},
        ]

        if format_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="candidate_import_template.csv"'
            writer = csv.DictWriter(response, fieldnames=['name', 'post', 'bio', 'image_url'])
            writer.writeheader()
            writer.writerows(sample)
            return response

        # --- Excel with field rules ---
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Protection, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        assert ws is not None  # wb.active is always set on a fresh Workbook()
        ws.title = 'Candidates'

        headers = ['name', 'post', 'bio', 'image_url']
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='FF9500', end_color='FF9500', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.protection = Protection(locked=True)

        # Sample rows
        for row_idx, s in enumerate(sample, 2):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=s[key])
                cell.font = Font(name='Calibri', size=11)
                cell.protection = Protection(locked=False)
                cell.border = thin_border

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 40

        # Data validation: post column → dropdown of actual position names
        if posts:
            post_list = ','.join(posts[:50])  # Excel DV limit
            post_dv = DataValidation(
                type='list',
                formula1=f'"{post_list}"',
                allow_blank=False,
                showErrorMessage=True,
                errorTitle='Invalid position',
                error='Position must match an existing position name.',
                showInputMessage=True,
                promptTitle='Position',
                prompt='Select the position this candidate is running for.',
            )
            post_dv.add('B2:B500')
            ws.add_data_validation(post_dv)

        # Protection
        ws.protection.sheet = True
        ws.protection.password = ''
        ws.protection.enable()
        for row in range(2, 501):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).protection = Protection(locked=False)
        ws.freeze_panes = 'A2'

        # Instructions sheet
        ins = wb.create_sheet('Instructions')
        ins_data = [
            ['Column', 'Description', 'Rules'],
            ['name', "Candidate's full name", 'Required. Max 255 characters.'],
            ['post', 'Position name', 'Required. Must match an existing position (use dropdown).'],
            ['bio', 'Short biography', 'Optional. Max 500 characters.'],
            ['image_url', 'Public URL of candidate photo', 'Optional. Must be https://. JPEG, PNG, GIF, or WebP. Max 5 MB.'],
        ]
        for row_idx, row_data in enumerate(ins_data, 1):
            for col_idx, val in enumerate(row_data, 1):
                cell = ins.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 1:
                    cell.font = Font(name='Calibri', bold=True, size=11)
        ins.column_dimensions['A'].width = 14
        ins.column_dimensions['B'].width = 35
        ins.column_dimensions['C'].width = 62
        ins.protection.sheet = True
        ins.protection.enable()

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="candidate_import_template.xlsx"'
        return response

    # ------------------------------------------------------------------
    # SSRF-safe image fetching
    # ------------------------------------------------------------------

    @staticmethod
    def _resolve_and_check(hostname: str):
        """Resolve hostname to a safe IP. Returns (ip_str, family) or (None, None) if blocked."""
        import socket
        try:
            results = socket.getaddrinfo(hostname, None, socket.AF_UNSPEC, socket.SOCK_STREAM)
            for family, _, _, _, sockaddr in results:
                ip = ipaddress.ip_address(sockaddr[0])
                blocked = any(ip in net for net in _BLOCKED_NETWORKS)
                if not blocked:
                    return sockaddr[0], family
            # All resolved IPs are private
            return None, None
        except (socket.gaierror, ValueError):
            return None, None

    @classmethod
    def _fetch_image_safely(cls, url: str):
        """
        Fetch an image from a URL with SSRF prevention.
        Returns a Django ContentFile or None.

        BE-40: DNS is resolved once and the URL is rewritten to use the
        resolved IP directly (thread-safe, no global monkey-patching).
        Redirects are followed manually with SSRF checks on every hop.
        """
        import socket
        import requests as req_lib
        from django.core.files.base import ContentFile

        # Validate URL scheme
        parsed = urlparse(url)
        if parsed.scheme not in ('http', 'https'):
            logger.warning("Blocked non-HTTP image URL: %s", url[:200])
            return None

        hostname = parsed.hostname or ''
        if not hostname:
            return None

        # BE-40: Resolve once, pin connection to the resolved IP
        resolved_ip, _family = cls._resolve_and_check(hostname)
        if resolved_ip is None:
            logger.warning("Blocked private/loopback image URL: %s", url[:200])
            return None

        # ------------------------------------------------------------------
        # Thread-safe DNS pinning: rewrite URL to use pre-resolved IP
        # directly instead of monkey-patching urllib3's create_connection.
        # The Host header preserves the original hostname for TLS/SNI.
        # ------------------------------------------------------------------
        parsed = urlparse(url)
        pinned_url = urlunparse(parsed._replace(netloc=f"{resolved_ip}:{parsed.port or (443 if parsed.scheme == 'https' else 80)}"))

        session = req_lib.Session()
        try:
            resp = session.get(
                pinned_url,
                timeout=_IMAGE_FETCH_TIMEOUT,
                stream=True,
                allow_redirects=False,
                headers={
                    'User-Agent': 'ElectON/2.0 ImageFetcher',
                    'Host': hostname,
                },
                verify=True,
            )
        except req_lib.RequestException as exc:
            logger.warning("Image fetch failed for %s: %s", url[:200], exc)
            return None

        pin_map: dict[str, str] = {hostname: str(resolved_ip)}
        try:
            # Follow redirects manually with SSRF check on each hop
            redirect_count = 0
            while resp.is_redirect and redirect_count < 5:
                redirect_count += 1
                redirect_url = resp.headers.get('Location', '')
                rp = urlparse(redirect_url)
                rhost = rp.hostname or ''
                if rhost and rhost != hostname:
                    rip, _ = cls._resolve_and_check(rhost)
                    if rip is None:
                        logger.warning("Redirect to private IP blocked: %s → %s", url[:200], redirect_url[:200])
                        return None
                    # Pin the new host for subsequent connections
                    pin_map[rhost] = str(rip)
                try:
                    resp = session.get(
                        redirect_url,
                        timeout=_IMAGE_FETCH_TIMEOUT,
                        stream=True,
                        allow_redirects=False,
                        headers={'User-Agent': 'ElectON/2.0 ImageFetcher'},
                    )
                except req_lib.RequestException as exc:
                    logger.warning("Image redirect fetch failed: %s: %s", redirect_url[:200], exc)
                    return None

            # BUG-3 fix: If we exhausted redirects, bail out explicitly
            if resp.is_redirect:
                logger.warning("Max redirects exceeded for image URL: %s", url[:200])
                return None

            resp.raise_for_status()

            # Check content type
            ct = resp.headers.get('Content-Type', '').split(';')[0].strip().lower()
            if ct not in _ALLOWED_IMAGE_CONTENT_TYPES:
                logger.warning("Image URL returned invalid content-type %s: %s", ct, url[:200])
                return None

            # Check content length header (if present)
            cl = resp.headers.get('Content-Length')
            if cl and int(cl) > _IMAGE_FETCH_MAX_SIZE:
                logger.warning("Image URL exceeds size limit (%s bytes): %s", cl, url[:200])
                return None

            # Read with size cap
            chunks = []
            total = 0
            for chunk in resp.iter_content(8192):
                total += len(chunk)
                if total > _IMAGE_FETCH_MAX_SIZE:
                    logger.warning("Image URL exceeded size limit during download: %s", url[:200])
                    return None
                chunks.append(chunk)

            data = b''.join(chunks)
            if not data:
                return None

            # BE-25: Validate magic bytes before Pillow verify
            _MAGIC_BYTES = {
                b'\xff\xd8\xff': 'image/jpeg',
                b'\x89PNG\r\n\x1a\n': 'image/png',
                b'GIF87a': 'image/gif',
                b'GIF89a': 'image/gif',
                b'BM': 'image/bmp',
            }
            magic_ok = any(data.startswith(sig) for sig in _MAGIC_BYTES)
            # LOW-23: Tighten RIFF check — verify WEBP subtype at offset 8
            if not magic_ok and data[:4] == b'RIFF' and data[8:12] == b'WEBP':
                magic_ok = True
            if not magic_ok:
                logger.warning("Image URL magic bytes don't match any known image format: %s", url[:200])
                return None

            # BE-41: Use img.load() instead of img.verify() to decompress
            # and catch decompression bombs / truncated files
            from PIL import Image as PILImage
            try:
                img = PILImage.open(io.BytesIO(data))
                img.load()  # Actually decompress pixel data
            except Exception:
                logger.warning("Image URL did not contain valid image data: %s", url[:200])
                return None

            # Derive filename from URL
            path = parsed.path.rstrip('/')
            filename = os.path.basename(path) or 'image.jpg'
            # Sanitize filename
            filename = re.sub(r'[^\w.\-]', '_', filename)[:100]

            return ContentFile(data, name=filename)

        except req_lib.RequestException as e:
            logger.warning("Failed to fetch image URL %s: %s", url[:200], str(e)[:200])
            return None
        except Exception:
            logger.exception("Unexpected error fetching image URL: %s", url[:200])
            return None

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _read_csv(uploaded_file):
        """Read CSV, trying multiple encodings. Returns list[dict] or None."""
        raw = uploaded_file.read()
        uploaded_file.seek(0)

        for encoding in ('utf-8', 'utf-8-sig', 'latin-1', 'cp1252'):
            try:
                text = raw.decode(encoding)
                rows = list(csv.DictReader(io.StringIO(text)))
                return [
                    {
                        k.lower().strip().replace(' ', '_'): (v or '').strip()
                        for k, v in row.items()
                        if k
                    }
                    for row in rows
                    if any((v or '').strip() for v in row.values())
                ]
            except UnicodeDecodeError:
                continue
        return None

    @staticmethod
    def _read_excel(uploaded_file):
        """Read Excel file. Returns list[dict] or None."""
        try:
            df = pd.read_excel(uploaded_file, engine='openpyxl')
            normalized = []
            for _, row in df.iterrows():
                d = {}
                for col in df.columns:
                    key = str(col).lower().strip().replace(' ', '_')
                    val = row[col]
                    d[key] = '' if pd.isna(val) else str(val).strip()
                if any(d.values()):
                    normalized.append(d)
            return normalized
        except Exception:
            logger.exception("Failed to read Excel file")
            return None

    def _validate_voter_rows(self, data):
        """Return list of validated {voter_name, voter_email} dicts.

        Accepts both the canonical column names (voter_name / voter_email) AND
        the legacy aliases (name / email) so that files exported by older
        templates still import cleanly.
        voter_name is optional — voter_email is the only required column.
        """
        if not data:
            self.errors.append("No data found in file.")
            return []

        first_row = data[0]

        # Accept legacy 'name'/'email' columns in addition to voter_name/voter_email
        has_email_col = 'voter_email' in first_row or 'email' in first_row
        has_name_col  = 'voter_name' in first_row or 'name'  in first_row

        if not has_email_col:
            self.errors.append(
                "Missing required column: 'Voter Email' (or 'voter_email' / 'email'). "
                "Download the template from the import panel for the correct format."
            )
            return []

        seen_emails: set[str] = set()
        validated: list[dict] = []

        for idx, row in enumerate(data, start=1):
            # Resolve voter_name with fallback to legacy 'name' column
            if has_name_col:
                name = (row.get('voter_name') or row.get('name') or '').strip()[:_MAX_FIELD_LENGTH]
            else:
                name = ''

            # Resolve voter_email with fallback to legacy 'email' column
            email = (row.get('voter_email') or row.get('email') or '').strip().lower()[:254]

            if not email:
                self.errors.append(f"Row {idx}: voter_email is required.")
                continue

            try:
                validate_email(email)
            except ValidationError:
                self.errors.append(f"Row {idx}: invalid email '{email}'.")
                continue

            if email in seen_emails:
                self.errors.append(f"Row {idx}: duplicate email '{email}' within file.")
                continue

            seen_emails.add(email)
            validated.append({'voter_name': name, 'voter_email': email})

        return validated

    def _deduplicate_voters(self, validated, election):
        """Split validated rows into unique / already-existing."""
        # BE-38: Case-insensitive email dedup
        existing_emails = set(
            e.lower()
            for e in election.voter_credentials.values_list('voter_email', flat=True)
        )
        unique, dupes = [], []
        for row in validated:
            if row['voter_email'].lower() in existing_emails:
                dupes.append(row)
                self.warnings.append(
                    f"Email {row['voter_email']} already exists in this election."
                )
            else:
                unique.append(row)
        return unique, dupes

    def _sanitize_field(self, value: str, field_name: str, row_idx: int, required: bool = True):
        """Sanitize and validate a text field from import data."""
        value = (value or '').strip()[:_MAX_FIELD_LENGTH]
        # Strip null bytes and control characters
        value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)
        if required and not value:
            self.errors.append(f"Row {row_idx}: {field_name} is required.")
            return None
        return value

    @staticmethod
    def _fail(message, errors=None):
        return {
            'success': False,
            'message': message,
            'processed': 0,
            'errors': errors or [message],
        }
